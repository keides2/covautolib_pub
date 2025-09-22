#!/usr/bin/env python3
# coding: utf-8

import sys
import os
import glob
from datetime import datetime
import json
import csv
import requests
from requests.exceptions import Timeout, ConnectionError, HTTPError, TooManyRedirects

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
import urllib3
from urllib3.exceptions import InsecureRequestWarning
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from logging import getLogger, Formatter, StreamHandler, FileHandler, DEBUG
import pandas as pd

# パスワード生成
import secrets
import string


# InsecureRequestWarning: 対策
#  Unverified HTTPS request is being made to host '<proxy host>'.
#  Adding certificate verification is strongly advised.
#  See: https://urllib3.readthedocs.io/en/latest/advanced-usage.html
urllib3.disable_warnings(InsecureRequestWarning)


def _build_proxy_dict(http_value, https_value):
    """HTTP/HTTPS プロキシ設定の辞書を生成する"""
    proxies = {}

    if http_value:
        proxies["http"] = http_value

    if https_value:
        proxies["https"] = https_value

    return proxies or None


# デフォルトプロキシ設定
DEFAULT_HTTP_PROXY = os.getenv("COVAUTO_HTTP_PROXY")
DEFAULT_HTTPS_PROXY = os.getenv("COVAUTO_HTTPS_PROXY") or DEFAULT_HTTP_PROXY
DEFAULT_PROXIES = _build_proxy_dict(DEFAULT_HTTP_PROXY, DEFAULT_HTTPS_PROXY)
ALT_HTTP_PROXY = os.getenv("COVAUTO_ALT_HTTP_PROXY")
ALT_HTTPS_PROXY = os.getenv("COVAUTO_ALT_HTTPS_PROXY") or ALT_HTTP_PROXY
ALT_PROXIES = _build_proxy_dict(ALT_HTTP_PROXY, ALT_HTTPS_PROXY)
REQUESTS_PROXIES = ALT_PROXIES or DEFAULT_PROXIES

if DEFAULT_PROXIES:
    http_proxy = DEFAULT_PROXIES.get("http")
    https_proxy = DEFAULT_PROXIES.get("https")

    if http_proxy:
        os.environ.setdefault("http_proxy", http_proxy)

    if https_proxy:
        os.environ.setdefault("https_proxy", https_proxy)


# os.sep は、Windows: \、 Linux: /
sep = os.sep

# JSON クラス
class GLJson:
    """
    GLJson JSONクラス
    """

    # JSONファイルの読み込み
    def read_json(self, file_path, encoding):
        """
        機能: JSON形式のファイルを読み、結果を辞書にして返す
        入力: 読み出しファイル名（パス付き）、
        エンコード形式（BOM有りUTF-8: "utf-8"、Shift-JIS: "shift_jis" は UnicodeDecodeError: 'shift_jis' codec can't decode）
        出力（戻り値）: JSON形式の辞書
        """

        try:
            f = open(file_path, "r", encoding=encoding)
        except FileNotFoundError as err:
            print("FileNotFoundError!: ", err)
            sys.exit(1)
        else:
            dict = json.load(f)
        f.close

        return dict

    # JSONファイルの読み込み２
    def read_json_2(self, file_path, encoding):
        """
        機能: JSON形式のファイルを読み、結果を辞書にして返す
        入力: 読み出しファイル名（パス付き）、
        エンコード形式（BOM有りUTF-8: "utf-8"、Shift-JIS: "shift_jis" は UnicodeDecodeError: 'shift_jis' codec can't decode）
        出力（戻り値）: 異常コード err, JSON形式の辞書
        """

        err = 0
        dict = {}

        try:
            f = open(file_path, "r", encoding=encoding)
            dict = json.load(f)

        except FileNotFoundError as e:
            print("Woops!: {}".format(e))

            err = 500
            return err, dict

        except UnicodeDecodeError as e:
            print("Woops!: {}".format(e))

            err = 501
            return err, dict

        f.close

        return err, dict

    # JSONファイルの保存1
    def write_json(self, file_path, dict, urlencode):
        """
        機能: JSON形式の辞書をBOM有りUTF-8にエンコードしてファイルに書き出す
        入力: 書き出しファイル名、書き出す辞書、URLエンコードの有無（する=True、しない=False）
        出力: 書き出しファイル
        戻り値: なし
        """

        f = open(file_path, "w", encoding="utf-8")
        # urlencode: True=URLエンコードする, False=URLエンコードしない、インデント=4、改行あり
        json.dump(dict, f, ensure_ascii=urlencode, indent=4)
        # print("write_json() 戻り値: ", j, type(j))  # None <class 'NoneType'>
        f.close

    # JSONファイルの保存2
    def write_json2(self, file_path, dict, encoding, urlencode):
        """
        機能: JSON形式の辞書をエンコード形式（BOM有りUTF-8: "utf-8"、Shift-JIS: "shift_jis"）を指定してファイルに書き出す
        入力: 書き出しファイル名、書き出す辞書、エンコード形式、URLエンコードの有無（する=True、しない=False）
        出力: 書き出しファイル
        戻り値: なし
        """

        f = open(file_path, "w", encoding=encoding)
        # urlencode: True=URLエンコードする, False=URLエンコードしない、インデント=4、改行あり
        json.dump(dict, f, ensure_ascii=urlencode, indent=4)
        # print("write_json() 戻り値: ", j, type(j))  # None <class 'NoneType'>
        f.close

    # 文字列が有効な JSON であるかどうかを確認する
    def is_json(self, myjson):
        """
        機能: 文字列が有効な JSON形式であるかどうかを確認する
        入力: JSON形式の辞書
        出力: 結果（True: 有効なJSON、False: 無効なJSON）
        """

        try:
            # 文字列型→辞書
            json_object = json.loads(myjson)
        except ValueError as e:
            print(e)
            return False

        return True

    # 日付が最新のファイルを取得
    def get_latest_json(self, file_path):
        """
        特定フォルダ内で日付が最新のファイルを取得
        入力: パスを含むファイル名 e.g.: DEST2 + "*" + os.sep + "*.json"
        戻り値: 最新ファイル名
        """
        list_of_files = glob.glob(file_path)
        # print("list_of_files: {}\n".format(list_of_files))
        latest_file = max(list_of_files, key=os.path.getctime)
        print("latest_file: {}\n".format(latest_file))

        return latest_file


# Excel クラス
class GLExcel:
    """
    GLExcel エクセル・クラス
    """

    # コンストラクタ
    def __init__(self, xlsx_path):
        """
        機能: コンストラクタ
        入力: ファイルパス
        出力: なし
        """
        self.xlsx_path = xlsx_path

    # エクセルファイルから読み取り
    def read_xlsx(self, sheet):
        """
        機能: CNの値をエクセルファイルから読み取る
        入力: シート名
        出力（戻り値）: CNのリスト
        """

        # workbookオブジェクトを作成
        wb = openpyxl.load_workbook(self.xlsx_path)
        # worksheetオブジェクトを作成
        wb_sheet = wb[sheet]

        # CNリスト e.g.: abeso, aihoshi, ...
        cn_list = []

        # last_row, last_column にワークシートの最終行と列を取得
        last_row = wb_sheet.max_row  # 最終行
        last_column = wb_sheet.max_column  # 最終列
        print("read_xlsx(): ", "最終行: ", last_row, "最終列: ", last_column)

        # A列 2行目から最終行目まで
        for row in range(2, last_row + 1):
            cell_value = wb_sheet["A" + str(row)].value
            cn_list.append(cell_value)

        print("cn_list: ", cn_list)
        return cn_list

    # エクセルファイルに記入
    def write_xlsx(self, group, cdate_start_end, dict):
        """
        機能: 抽出結果をエクセルファイルに記入する
        入力: エクセルファイルパス、グループ名、調査開始非―調査終了日、抽出結果辞書
        出力: 抽出結果が追記されたエクセルファイル
        """

        # workbookオブジェクトを作成
        wb = openpyxl.load_workbook(self.xlsx_path)
        # worksheetオブジェクトを作成
        wb_sheet = wb["Sheet1"]

        # last_column にワークシートの最終列を取得
        # last_column = wb_sheet.max_column

        # 調査開始日セット
        wb_sheet["T1"] = cdate_start_end

        # last_row, last_column にワークシートの最終行と列を取得
        last_row = wb_sheet.max_row  # 最終行
        last_column = wb_sheet.max_column  # 最終列
        print("write_xlsx(): ", "最終行: ", last_row, "最終列: ", last_column)

        count = dict["summary"]["count"]
        cves = str(dict["summary"]["cves"])

        # 角括弧削除
        cves = cves.replace("[", "")
        cves = cves.replace("]", "")

        # JSON→文字列型
        dict_json = json.dumps(dict)
        print(GLJson.is_json(dict_json))  # True?

        # D列 11行目から最終行目まで
        for row in range(11, last_row + 1):
            cell_value = wb_sheet["D" + str(row)].value
            print("group: ", group, " ,row: ", row, "cell_value: ", cell_value)
            if cell_value == group:
                # print("match")
                wb_sheet["T" + str(row)] = count
                if cves != "[]":
                    # 正規表現で抽出（'[', ']'除去）
                    # cves = re.findall("[(.*)]", cves)
                    wb_sheet["U" + str(row)] = cves
                else:
                    wb_sheet["U" + str(row)] = "-"
                break

        # 上書き保存
        wb.save(self.xlsx_path)

    # エクセルファイルの前処理
    def prepare_xlsx(self):
        """
        機能: 抽出結果をエクセルファイルに追記する前に、T列・U列を追加、罫線の描画等を行う
        入力: 抽出結果エクセルファイルパス
        出力: T列・U列を追加、罫線の描画等した抽出結果エクセルファイル
        """

        # エクセルファイルを開いて、T列から2列追加
        wb = openpyxl.load_workbook(self.xlsx_path)
        # worksheet オブジェクトを作成
        wb_sheet = wb["Sheet1"]

        # last_row, last_column にワークシートの最終行と列を取得
        last_row = wb_sheet.max_row  # 最終行
        last_column = wb_sheet.max_column  # 最終列
        print("prepare_xlsx(): ", "最終行: ", last_row, "最終列: ", last_column)

        # T列から2列追加
        wb_sheet.insert_cols(20, 2)

        # 調査開始日とCVE追記
        wb_sheet["T2"] = "件数"
        wb_sheet["U2"] = "CVE"

        # 複数セルを太字
        font = Font(bold=True)
        for row_num in range(1, 3):  # 1, 2行目
            for col_num in range(20, 22):  # T, U列 = 20, 21
                wb_sheet.cell(row=row_num, column=col_num).font = font

        # 普通線、色は自動
        side = Side(style="medium")
        border = Border(top=side, bottom=side, left=side, right=side)

        # 複数セルに罫線を引く
        for row_num in range(2, last_row + 1):  # 2行目から
            for col_num in range(20, 22):  # T列 = 20 から 2列
                wb_sheet.cell(row=row_num, column=col_num).border = border

        # 上書き保存
        wb.save(self.xlsx_path)

    # GitLab ユーザー追加
    def gl_add_users(self, list):
        """
        機能: エクセルファイルに新しいユーザーを追記する
        入力: 追記するユーザーリスト
        出力: 更新されたエクセルファイル
        """

        # エクセルファイルを開く
        wb = openpyxl.load_workbook(self.xlsx_path)
        # worksheet オブジェクトを作成
        wb_sheet = wb["GitLab"]

        # last_row, last_column にワークシートの最終行と列を取得
        last_row = wb_sheet.max_row  # 最終行
        last_column = wb_sheet.max_column  # 最終列
        print("gl_add_users(): ", "最終行: ", last_row, "最終列: ", last_column)

        # エクセルに追記
        # テーブルのスタイルに書式設定をしてはいけない（エラーになる）
        len_add = len(list)
        for n in range(len_add):
            # id 列に id の値を追加
            wb_sheet["A" + str(last_row + len_add - n)] = list[n]["id"]

            # username 列に username の値を追加
            wb_sheet["B" + str(last_row + len_add - n)] = list[n]["username"]

            # name 列に name の値を追加
            wb_sheet["D" + str(last_row + len_add - n)] = list[n]["name"]

            # state 列に state の値を追加
            wb_sheet["G" + str(last_row + len_add - n)] = list[n]["state"]

            # avatar_url 列に avatar_url の値を追加
            if list[n]["avatar_url"]:
                wb_sheet["H" + str(last_row + len_add - n)] = list[n]["avatar_url"]
            else:
                pass

            # web_url 列に web_url の値を追加
            wb_sheet["I" + str(last_row + len_add - n)] = list[n]["web_url"]

        # 上書き保存
        wb.save(self.xlsx_path)

    # GitLab ユーザーの id_max を求める
    def get_id_max(self):
        """
        機能: 最終行のA列から id の最大値を求める
        入力: なし
        出力: id_max
        """

        # エクセルファイルを開く
        wb = openpyxl.load_workbook(self.xlsx_path)
        # worksheet オブジェクトを作成
        wb_sheet = wb["GitLab"]

        # last_row, last_column にワークシートの最終行と列を取得
        last_row = wb_sheet.max_row  # 最終行
        last_column = wb_sheet.max_column  # 最終列
        print("gl_add_users(): ", "最終行: ", last_row, "最終列: ", last_column)

        id_max = wb_sheet["A" + str(last_row)].value

        return id_max


# メール（送信のみ）クラス
class GLMail:
    """
    GLMail メール（送信のみ）クラス
    """

    # メール送信コンストラクタ
    def __init__(self, from_address, to_address, cc_address, subject, body, file_list):
        """
        機能: メールを送信する（コンストラクタ）
        入力: From、To、Cc、Subject、本文、添付するファイルのリスト（フォルダのパス付）
        出力: メール送信
        """
        self.from_address = from_address
        self.to_address = to_address
        self.cc_address = cc_address
        self.subject = subject
        self.body = body
        self.file_list = file_list

    # メール送信
    def send_mail(self):
        """
        機能: メールを送信する
        入力: なし
        出力: なし（送信メール）
        戻り値: 正常終了/異常終了(0/1)
        """

        TO_CC_ADDRESS = self.to_address + self.cc_address
        print(TO_CC_ADDRESS, type(TO_CC_ADDRESS))
        smtp_host = os.getenv("COVAUTO_SMTP_HOST", "localhost")
        smtp_port = int(os.getenv("COVAUTO_SMTP_PORT", "25"))

        msg = MIMEMultipart("alternative")  # HTML
        msg["Subject"] = self.subject
        msg["From"] = self.from_address
        msg["To"] = ", ".join(self.to_address)
        msg["CC"] = ", ".join(self.cc_address)
        msg.attach(MIMEText(self.body, "html"))

        # 複数のファイルを添付する
        for file in self.file_list:
            fname = os.path.basename(file)
            part = MIMEBase("application", "octet-stream")
            part.set_payload(open(file, "rb").read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=fname)
            msg.attach(part)

        try:
            smtpobj = smtplib.SMTP(smtp_host, smtp_port)
            smtpobj.sendmail(self.from_address, TO_CC_ADDRESS, msg.as_string())
            smtpobj.close()

            return 0

        except Exception as e:
            print("SMTP error!: {}".format(e))

            return 1


# プロジェクト・クラス
class COVProj:
    """
    COVProj プロジェクト・クラス
    実行環境を判定し、環境を整える
    """

    # コンストラクタ
    def __init__(self, group, project, branch, commit):
        """
        機能: コンストラクタ
        入力: グループ名、プロジェクト名、ブランチ名、コミットメッセージ
        出力: インスタンス内グループ名、プロジェクト名、ブランチ名、コミットメッセージ、
              ディレクトリ名（基底、共有、コンフィグ、アドレス、認証キー）
              ※1 認証キーは、共有ディレクトリに置いてはいけない（Sドライブは所有者をvulsから変更できないので置けない）
        """
        # 環境判断
        # 実行環境に応じてディレクトリとプロキシを初期化する
        sep = os.sep

        if os.name == "nt":
            default_base_dir = os.path.join(os.path.expanduser("~"), "cov")
        else:
            default_base_dir = "/cov"

        base_dir = os.getenv("COVAUTO_BASE_DIR", default_base_dir)
        base_dir = os.path.abspath(base_dir)
        if not base_dir.endswith(sep):
            base_dir += sep

        shared_dir_env = os.getenv("COVAUTO_SHARE_DIR")
        share_dir = os.path.abspath(shared_dir_env) if shared_dir_env else base_dir
        if not share_dir.endswith(sep):
            share_dir += sep

        runtime_http_proxy = os.getenv("COVAUTO_RUNTIME_HTTP_PROXY")
        runtime_https_proxy = os.getenv("COVAUTO_RUNTIME_HTTPS_PROXY") or runtime_http_proxy
        runtime_proxies = _build_proxy_dict(runtime_http_proxy, runtime_https_proxy) or REQUESTS_PROXIES

        if runtime_proxies:
            http_proxy = runtime_proxies.get("http")
            https_proxy = runtime_proxies.get("https")
            if http_proxy:
                os.environ["http_proxy"] = http_proxy
            if https_proxy:
                os.environ["https_proxy"] = https_proxy

        # グループ・ディレクトリが存在しないなら作成
        groups_dir = base_dir + "groups"
        if not os.path.isdir(groups_dir):
            os.mkdir(groups_dir)

        # 基底ディレクトリ
        self.base_dir = groups_dir + sep
        # 共有ディレクトリ
        self.share_dir = share_dir
        # コンフィグ・ディレクトリ
        self.cfg_dir = share_dir + "config" + sep
        # アドレス・ディレクトリ
        self.address_dir = share_dir + "address" + sep
        # 認証キー・ディレクトリ
        self.auth_key_dir = base_dir + "auth-key" + sep  # ※1

        # log ディレクトリが存在しないなら作成
        log_dir = base_dir + "log" + sep
        if not os.path.isdir(log_dir):
            os.mkdir(log_dir)

        # log ディレクトリ
        self.log_dir = log_dir

        # プロジェクトまでのパス
        self.path_to_project = self.base_dir + group + sep + project + sep

        # ブランチまでのパス
        self.path_to_branch = self.base_dir + group + sep + project + sep + branch + sep

        # インスタンス変数
        self.group = group
        self.project = project
        self.branch = branch
        self.commit = commit
        self.commit_short = commit[
            :20
        ]  # コミット・メッセージ（先頭から20文字の短縮型）


# Logger クラス
class LOGGER:
    """
    機能: ログ採取（画面とファイルの両方に出力可能）
    入力: ログファイルのヘッダ名、COVProj クラスのインスタンス
    出力: ファイル名: header_group_project_branch-YYYYMMDDhhmmss.log
        (例) covauto_user2_cov_auto_main-20221228114113.log
        形式:
        2022-12-28 11:41:13 - covauto - DEBUG - [config] COV-CONFIGURE_OP: --python --version 3
        2022-12-28 11:41:13 - covauto - ERROR - [GIT] git fetch で ...
        2022-12-28 11:41:13 - covauto - INFO - [main] CONFIG ...
        2022-12-28 11:41:13 - covauto - WARNING - [error_proc] error_proc が呼ばれ ...
    """

    def __init__(self, header, covproj):
        # logger
        # ロガー・オブジェクトの生成
        # self.logger = getLogger('covlog')
        self.logger = getLogger(header)

        # ロガーの出力レベル設定(ハンドラに渡すエラーメッセージの最低レベル)
        self.logger.setLevel(DEBUG)

        # 今日の日付を取得
        dt_today = datetime.today()
        str_month = format(str(dt_today.month), "0>2")  # 2桁、ゼロ埋め
        str_day = format(str(dt_today.day), "0>2")  # 2桁、ゼロ埋め
        str_today = str(dt_today.year) + str_month + str_day
        # 今の時刻を取得
        str_hour = format(str(dt_today.hour), "0>2")  # 2桁、ゼロ埋め
        str_minute = format(str(dt_today.minute), "0>2")  # 2桁、ゼロ埋め
        str_second = format(str(dt_today.second), "0>2")  # 2桁、ゼロ埋め
        str_time = str_hour + str_minute + str_second

        # ハンドラーの生成
        stream_handler = StreamHandler()
        filename = (
            covproj.log_dir
            + header
            + "_"
            + covproj.group
            + "_"
            + covproj.project
            + "_"
            + covproj.branch
            + "-"
            + str_today
            + str_time
            + ".log"
        )
        file_handler = FileHandler(filename)

        # ハンドラーのログレベル設定(ハンドラが出力するエラーメッセージの最低レベル)
        stream_handler.setLevel(DEBUG)
        file_handler.setLevel(DEBUG)

        # ログ出力フォーマット設定
        handler_format = Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        stream_handler.setFormatter(handler_format)
        file_handler.setFormatter(handler_format)

        # ロガーにハンドラーをセット
        self.logger.addHandler(stream_handler)
        self.logger.addHandler(file_handler)

        # ロガーのインスタンス変数にファイル名をセット
        self.logger.filename = filename


# Logger_2 クラス
class LOGGER_2:
    """
    機能: ログ採取（画面とファイルの両方に出力可能）
    入力: ログファイルのヘッダ名、スナップショットID
    出力: ファイル名: header_snapshot_id-YYYYMMDDhhmmss.log
        (例) cov_snap_14185-20230515114113.log
        形式:
        2022-12-28 11:41:13 - cov_snap - DEBUG - [config] COV-CONFIGURE_OP: --python --version 3
        2022-12-28 11:41:13 - cov_snap - ERROR - [GIT] git fetch で ...
        2022-12-28 11:41:13 - cov_snap - INFO - [main] CONFIG ...
        2022-12-28 11:41:13 - cov_snap - WARNING - [error_proc] error_proc が呼ばれ ...
    """

    def __init__(self, header, snapshot_id):
        # logger
        # ロガー・オブジェクトの生成
        # self.logger = getLogger('covlog')
        self.logger = getLogger(header)

        # ロガーの出力レベル設定(ハンドラに渡すエラーメッセージの最低レベル)
        self.logger.setLevel(DEBUG)

        # 今日の日付を取得
        dt_today = datetime.today()
        str_month = format(str(dt_today.month), "0>2")  # 2桁、ゼロ埋め
        str_day = format(str(dt_today.day), "0>2")  # 2桁、ゼロ埋め
        str_today = str(dt_today.year) + str_month + str_day
        # 今の時刻を取得
        str_hour = format(str(dt_today.hour), "0>2")  # 2桁、ゼロ埋め
        str_minute = format(str(dt_today.minute), "0>2")  # 2桁、ゼロ埋め
        str_second = format(str(dt_today.second), "0>2")  # 2桁、ゼロ埋め
        str_time = str_hour + str_minute + str_second

        # ハンドラーの生成
        stream_handler = StreamHandler()
        filename = (
            "."
            + sep
            + "log"
            + sep
            + header
            + "_"
            + snapshot_id
            + "-"
            + str_today
            + str_time
            + ".log"
        )
        file_handler = FileHandler(filename)

        # ハンドラーのログレベル設定(ハンドラが出力するエラーメッセージの最低レベル)
        stream_handler.setLevel(DEBUG)
        file_handler.setLevel(DEBUG)

        # ログ出力フォーマット設定
        handler_format = Formatter(
            "%(asctime)s - %(name)s - %(levelname)s - %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        stream_handler.setFormatter(handler_format)
        file_handler.setFormatter(handler_format)

        # ロガーにハンドラーをセット
        self.logger.addHandler(stream_handler)
        self.logger.addHandler(file_handler)

        # ロガーのインスタンス変数にファイル名をセット
        self.logger.filename = filename


# Coverity API クラス
class COVApi:
    """
    COVApi APIクラス
    """

    # コンストラクタ
    def __init__(self, group=None, project=None, branch=None, commit=None):
        """
        引数のデフォルト値をNoneに設定し、複数のコンストラクタに対応する
        """
        # [API reference Swagger のページ](https://<coverity-host>/swagger/api/v2/index.html)

        # API base url
        self.API_BASE = os.getenv("COVAUTO_API_BASE_URL", "https://<coverity-host>/api/v2")

        # GET /issues/columns
        # API の分類 Issues から GET /issues/columns を使い CID、CategoryとImpact列の列キーを取得する
        self.API_ISSUES_COLUMNS = "/issues/columns"

        # GET /projects
        # API の分類 Project から GET /projects を使い、すべてのプロジェクト名の取得、出力ファイル projects.json からプロジェクト名、プロジェクトキー、ストリーム名、を取得する
        self.API_PROJECTS = "/projects"

        # GET /streams
        # API の分類 Streams から GET /streams を使い、すべてのストリーム名を取得する
        self.API_STREAMS = "/streams"

        # GET /streams/stream/snapshots
        # API の分類 Streams から GET /streams/stream/snapshots を使い、スナップショットIDを取得する
        self.API_SNAPSHOTS = "/streams/stream/snapshots"

        # GET /views
        # API の分類 Views から GET /views を使い、現在のユーザーがアクセスできるすべてのビューを取得する
        self.API_VIEWS = "/views"

        # GET /views/user
        # API の分類 Views から GET /views/user を使い、ビュー「無条件全表示のコピー」の ID を取得する
        self.API_VIEWS_USER = "/views/user"

        #  GET /views/viewContents/{id}
        # API の分類 Views から GET /views/viewContents/{id} を使い、ビューの項目と指摘内容を取得する
        self.API_VIEWS_VIEWCONTENTS = "/views/viewContents"

        # POST /issues/search
        # API の分類 Issues から POST /issues/search を使い指摘内容を取得する
        self.API_POST_ISSUES_SEARCH = "/issues/search"

        # POST /issueOccurrencs/search
        # API の分類 Issue Occurrences から POST /issueOccurrencs/search を使い指摘内容を取得する
        self.API_POST_ISSUEOCCURRENCES_SEARCH = "/issueOccurrences/search"

        # 認証キー
        # self.API_AUTH = ("user2", "08EEDxxxxxxx2B6")
        username = self.get_env_variable("COVAUTHUSER")
        auth_key = self.get_env_variable("COVAUTHKEY")
        self.API_AUTH = (username, auth_key)

        # Get /users, /users/{name}
        self.api_userS = "/users"

        # チェッカー名一覧の取得（key, value のみ）
        self.API_CHECKER_ATTRIBUTES    = "/checkerAttributes"

        # headers
        self.API_HEADERS = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36 Edg/109.0.1518.61",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

        # インスタンス変数
        if group is not None:
            self.group = group
        else:
            self.group = "None"

        if project is not None:
            self.project = project
        else:
            self.project = "None"

        if branch is not None:
            self.branch = branch
        else:
            self.branch = "None"

        if commit is not None:
            self.commit = commit
            self.commit_short = commit[
                :20
            ]  # コミット・メッセージ（先頭から20文字の短縮型）
        else:
            self.commit = "None"
            self.commit_short = "None"

        print(
            "self.group: {}, self.project: {}, self.branch: {}, self.commit: {}, self.commit_short: {}".format(
                self.group, self.project, self.branch, self.commit, self.commit_short
            )
        )

    # GET /issues/columns
    def cov_get_issues_columns(self):
        """
        機能: GET /issues/columns を呼び出して列の名前 columnKey 一覧を取得する
        入力: なし
        出力: ファイル issues_columns.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/issues/columns?locale=ja_JP&queryType=bySnapshot&retrieveGroupByColumns=false" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_ISSUES_COLUMNS
        payload = {
            "locale": "ja_JP",
            "queryType": "bySnapshot",
            "retrieveGroupByColumns": "false",
        }
        print("cov_get_issues_columns:")
        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.headers)
        """response.headers
        {
            "Date": "Wed, 25 Jan 2023 06:53:20 GMT",
            "Content-Type": "application/json;charset=UTF-8",
            "Transfer-Encoding": "chunked",
            "Connection": "keep-alive",
            "Set-Cookie": "COVJSESSIONID8080BV=6812C84BD585DB0AC66F0059F22245D1; Path=/; HttpOnly",
            "Pragma": "no-cache",
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
            "X-Frame-Options": "DENY",
            "X-XSS-Protection": "1; mode=block",
            "X-Content-Type-Options": "nosniff",
            "vary": "accept-encoding",
            "Content-Encoding": "gzip",
            "Server": "Synopsys"
        }
        """
        # print(response.text)
        """response.text
        [
          {
            "columnKey": "cid",
            "name": "CID"
          },
          {
            "columnKey": "checker",
            "name": "Checker"
          },
          {
            "columnKey": "displayImpact",
            "name": "Impact"
          },
          {
            "columnKey": "displayCategory",
            "name": "Category"
          },
          {
            "columnKey": "displayType",
            "name": "Type"
          },
          {
            "columnKey": "cwe",
            "name": "CWE"
          },
          {
            "columnKey": "displayIssueKind",
            "name": "Issue Kind"
          },
          {
            "columnKey": "status",
            "name": "Status"
          },
          {
            "columnKey": "firstDetected",
            "name": "First Detected"
          },
          {
            "columnKey": "owner",
            "name": "Owner"
          },
          {
            "columnKey": "ownerFullName",
            "name": "Owner Name"
          },
          {
            "columnKey": "externalReference",
            "name": "External Reference"
          },
          {
            "columnKey": "classification",
            "name": "Classification"
          },
          {
            "columnKey": "severity",
            "name": "Severity"
          },
          {
            "columnKey": "action",
            "name": "Action"
          },
          {
            "columnKey": "fixTarget",
            "name": "Fix Target"
          },
          {
            "columnKey": "legacy",
            "name": "Legacy"
          },
          {
            "columnKey": "displayComponent",
            "name": "Component"
          },
          {
            "columnKey": "displayFile",
            "name": "File"
          },
          {
            "columnKey": "displayFunction",
            "name": "Function"
          },
          {
            "columnKey": "functionMergeName",
            "name": "Function Merge Name"
          },
          {
            "columnKey": "mergeExtra",
            "name": "Merge Extra"
          },
          {
            "columnKey": "mergeKey",
            "name": "Merge Key"
          },
          {
            "columnKey": "fileLanguage",
            "name": "Language"
          },
          {
            "columnKey": "lastTriaged",
            "name": "Last Triaged"
          },
          {
            "columnKey": "lastTriagedUser",
            "name": "Last Triaged User"
          },
          {
            "columnKey": "occurrenceCount",
            "name": "Count"
          },
          {
            "columnKey": "displayComparison",
            "name": "Comparison"
          },
          {
            "columnKey": "firstSnapshotDate",
            "name": "First Snapshot Date"
          },
          {
            "columnKey": "firstSnapshotId",
            "name": "First Snapshot"
          },
          {
            "columnKey": "firstSnapshotVersion",
            "name": "First Snapshot Version"
          },
          {
            "columnKey": "firstSnapshotTarget",
            "name": "First Snapshot Target"
          },
          {
            "columnKey": "firstSnapshotDescription",
            "name": "First Snapshot Description"
          },
          {
            "columnKey": "firstSnapshotStream",
            "name": "First Snapshot Stream"
          },
          {
            "columnKey": "lastDetected",
            "name": "Last Snapshot Date"
          },
          {
            "columnKey": "lastDetectedId",
            "name": "Last Snapshot"
          },
          {
            "columnKey": "lastDetectedVersion",
            "name": "Last Snapshot Version"
          },
          {
            "columnKey": "lastDetectedTarget",
            "name": "Last Snapshot Target"
          },
          {
            "columnKey": "lastDetectedDescription",
            "name": "Last Snapshot Description"
          },
          {
            "columnKey": "lastDetectedStream",
            "name": "Last Snapshot Stream"
          },
          {
            "columnKey": "score",
            "name": "Score"
          },
          {
            "columnKey": "lineNumber",
            "name": "Line Number"
          },
          {
            "columnKey": "lastTriageComment",
            "name": "Last Triage Comment"
          },
          {
            "columnKey": "column_standard_PCI DSS 2018",
            "name": "Standard: PCI DSS 2018"
          },
          {
            "columnKey": "column_standard_Payment Card Industry Data Security Standard (PCI DSS) 2018",
            "name": "Standard: Payment Card Industry Data Security Standard (PCI DSS) 2018"
          },
          {
            "columnKey": "column_standard_AUTOSAR C++14",
            "name": "Standard: AUTOSAR C++14"
          },
          {
            "columnKey": "column_standard_OWASP Web Top Ten 2021",
            "name": "Standard: OWASP Web Top Ten 2021"
          },
          {
            "columnKey": "column_standard_OWASP Web Top Ten 2017",
            "name": "Standard: OWASP Web Top Ten 2017"
          },
          {
            "columnKey": "column_standard_DISA-STIG V5",
            "name": "Standard: DISA-STIG V5"
          },
          {
            "columnKey": "column_standard_DISA-STIG V4R10",
            "name": "Standard: DISA-STIG V4R10"
          },
          {
            "columnKey": "column_standard_DISA-STIG V4R3",
            "name": "Standard: DISA-STIG V4R3"
          },
          {
            "columnKey": "column_standard_CERT C",
            "name": "Standard: CERT C"
          },
          {
            "columnKey": "column_standard_MISRA Category",
            "name": "Standard: MISRA Category"
          },
          {
            "columnKey": "column_standard_DISA-STIG Severity",
            "name": "Standard: DISA-STIG Severity"
          },
          {
            "columnKey": "column_standard_CERT C++",
            "name": "Standard: CERT C++"
          },
          {
            "columnKey": "column_standard_CERT JAVA",
            "name": "Standard: CERT JAVA"
          },
          {
            "columnKey": "column_standard_MISRA C 2012",
            "name": "Standard: MISRA C 2012"
          },
          {
            "columnKey": "column_standard_OWASP Mobile Top Ten 2016",
            "name": "Standard: OWASP Mobile Top Ten 2016"
          },
          {
            "columnKey": "column_standard_2021 CWE Top 25",
            "name": "Standard: 2021 CWE Top 25"
          },
          {
            "columnKey": "column_standard_MISRA C++ 2008",
            "name": "Standard: MISRA C++ 2008"
          },
          {
            "columnKey": "column_standard_MISRA C 2004",
            "name": "Standard: MISRA C 2004"
          },
          {
            "columnKey": "column_standard_DISA-STIG V5 Vulnerability ID (Vul ID)",
            "name": "Standard: DISA-STIG V5 Vulnerability ID (Vul ID)"
          },
          {
            "columnKey": "column_standard_DISA-STIG V4R10 Vulnerability ID (Vul ID)",
            "name": "Standard: DISA-STIG V4R10 Vulnerability ID (Vul ID)"
          },
          {
            "columnKey": "column_standard_ISO TS17961 2016",
            "name": "Standard: ISO TS17961 2016"
          },
          {
            "columnKey": "project",
            "name": "Project"
          }
        ]
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json[13], indent=4))
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        # インスタンス化しないと self.write_json2() で 'COVApi' object has no attribute 'write_json2'
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("issues_columns.json", res_json, "shift_jis", False)

    # GET /projects プロジェクト名からプロジェクトキー、ストリーム名を取得
    def cov_get_projects(self, project):
        """
        機能: GET /projects を呼び出して プロジェクト名からプロジェクトキー、ストリーム名を取得する
        入力: プロジェクト名
        出力: ファイル projects.json、戻り値 プロジェクトキー、ストリーム名
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/projects/sample_projectA?includeChildren=true&includeStreams=true&locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_PROJECTS + "/" + project
        payload = {
            "locale": "ja_JP",
            "includeChildren": "true",
            "includeStreams": "true",
        }
        print("cov_get_projects:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 07:35:39 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=341BC97D31ECB283FBD12E71236E25A7; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate, no-store",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Content-Language": "ja-JP",
          "Server": "Synopsys"
        }
        """
        # print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }
        """
        """response.text 成功の例
        {
          "projects": [
            {
              "createdBy": "user5",
              "dateCreated": "2022-12-27T06:32:28.278Z",
              "dateModified": "2022-12-27T06:32:28.278Z",
              "description": "sample_projectA",
              "modifiedBy": "user5",
              "name": "sample_projectA",
              "projectKey": 10072,
              "roleAssignments": [
                {
                  "group": null,
                  "roleAssignmentType": "user",
                  "roleName": "projectOwner",
                  "scope": "project",
                  "username": "user5"
                }
              ],
              "streamLinks": [],
              "streams": [
                {
                  "allowCommitWithoutPassword": null,
                  "analysisVersionOverride": null,
                  "autoDeleteOnExpiry": false,
                  "componentMapName": "Default",
                  "description": "sample_projectA",
                  "enableDesktopAnalysis": true,
                  "name": "sample_projectA_coverity",
                  "outdated": false,
                  "ownerAssignmentOption": "default_component_owner",
                  "pluginVersionOverride": null,
                  "primaryProjectName": "sample_projectA",
                  "roleAssignments": [],
                  "summaryExpirationDays": null,
                  "triageStoreName": "Default Triage Store",
                  "versionMismatchMessage": null
                }
              ]
            }
          ],
          "code": null,
          "message": null
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # projectKey 抽出
        projectKey = res_json["projects"][0]["projectKey"]
        # print("projectKey: {}".format(json.dumps(projectKey)))

        # ストリーム名の抽出
        stream = res_json["projects"][0]["streams"][0]["name"]
        # print("stream name: {}".format(json.dumps(stream)))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("projects.json", res_json, "shift_jis", False)

        return projectKey, stream

    # GET /projects すべてのプロジェクト名を取得
    def cov_get_all_projects(self):
        """
        機能: GET /projects を呼び出して すべてのプロジェクト名を取得する
        入力: なし
        出力: ファイル projects.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/projects?includeChildren=false&includeStreams=false&locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_PROJECTS
        payload = {
            "locale": "ja_JP",
            "includeChildren": "false",
            "includeStreams": "true",
        }
        print("cov_get_all_projects:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20
        )
        
        print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 07:35:39 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=341BC97D31ECB283FBD12E71236E25A7; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate, no-store",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Content-Language": "ja-JP",
          "Server": "Synopsys"
        }
        """

        print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }
        """
        """response.text 成功の例
        {
          "projects": [
            {
              "createdBy": "user5",
              "dateCreated": "2022-12-27T06:32:28.278Z",
              "dateModified": "2022-12-27T06:32:28.278Z",
              "description": "sample_projectA",
              "modifiedBy": "user5",
              "name": "sample_projectA",
              "projectKey": 10072,
              "roleAssignments": [
                {
                  "group": null,
                  "roleAssignmentType": "user",
                  "roleName": "projectOwner",
                  "scope": "project",
                  "username": "user5"
                }
              ],
              "streamLinks": [],
              "streams": [
                {
                  "allowCommitWithoutPassword": null,
                  "analysisVersionOverride": null,
                  "autoDeleteOnExpiry": false,
                  "componentMapName": "Default",
                  "description": "sample_projectA",
                  "enableDesktopAnalysis": true,
                  "name": "sample_projectA_coverity",
                  "outdated": false,
                  "ownerAssignmentOption": "default_component_owner",
                  "pluginVersionOverride": null,
                  "primaryProjectName": "sample_projectA",
                  "roleAssignments": [],
                  "summaryExpirationDays": null,
                  "triageStoreName": "Default Triage Store",
                  "versionMismatchMessage": null
                }
              ]
            }
          ],
          "code": null,
          "message": null
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("projects.json", res_json, "cp932", False)

        return res_json

    # GET /streams すべてのストリーム名を取得
    def cov_get_all_streams(self):
        """
        機能: GET /streams を呼び出して すべてのストリーム名を取得する
        入力: なし
        出力: ファイル streams.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/streams?=excludeRoles=true&locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_STREAMS
        payload = {
            "locale": "ja_JP",
            "excludeRoles": "true",
        }
        print("cov_get_all_streams:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20
        )
        
        print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 07:35:39 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=341BC97D31ECB283FBD12E71236E25A7; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate, no-store",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Content-Language": "ja-JP",
          "Server": "Synopsys"
        }
        """

        print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }
        """
        """response.text 成功の例
        {
        "streams": [
          {
            "allowCommitWithoutPassword": true,
            "analysisVersionOverride": "string",
            "autoDeleteOnExpiry": true,
            "componentMapName": "string",
            "description": "string",
            "enableDesktopAnalysis": true,
            "name": "string",
            "outdated": true,
            "ownerAssignmentOption": "default_component_owner",
            "pluginVersionOverride": "string",
            "primaryProjectName": "string",
            "roleAssignments": [
              {
                "group": {
                  "domainName": "string",
                  "ldapServer": "string",
                  "name": "string"
                },
                "roleAssignmentType": "group",
                "roleName": "string",
                "scope": "component",
                "username": "string"
              }
            ],
            "summaryExpirationDays": 0,
            "triageStoreName": "string",
            "versionMismatchMessage": "string"
          }
        ]
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("streams.json", res_json, "cp932", False)

        return res_json

    # GET /streams/stream/snapshots ストリーム内のすべてのスナップショットIDを取得
    def cov_get_all_snapshots_in_stream(self, stream):
        """
        機能: GET /streams/stream/snapshots を呼び出して ストリーム内のすべてのスナップショットIDを取得する
        入力: なし
        出力: ファイル snapshotsForStream.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/streams/stream/snapshots?idType=byName&name=My_Stream&locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_SNAPSHOTS
        payload = {
            "locale": "ja_JP",
            "idType": "byName",
            "name": stream
        }
        print("cov_get_all_snapshots_in_stream:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20
        )
        
        print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 07:35:39 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=341BC97D31ECB283FBD12E71236E25A7; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate, no-store",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Content-Language": "ja-JP",
          "Server": "Synopsys"
        }
        """

        print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }
        """
        """response.text 成功の例
        {
        "streams": [
          {
            "allowCommitWithoutPassword": true,
            "analysisVersionOverride": "string",
            "autoDeleteOnExpiry": true,
            "componentMapName": "string",
            "description": "string",
            "enableDesktopAnalysis": true,
            "name": "string",
            "outdated": true,
            "ownerAssignmentOption": "default_component_owner",
            "pluginVersionOverride": "string",
            "primaryProjectName": "string",
            "roleAssignments": [
              {
                "group": {
                  "domainName": "string",
                  "ldapServer": "string",
                  "name": "string"
                },
                "roleAssignmentType": "group",
                "roleName": "string",
                "scope": "component",
                "username": "string"
              }
            ],
            "summaryExpirationDays": 0,
            "triageStoreName": "string",
            "versionMismatchMessage": "string"
          }
        ]
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("snapshotsForStream.json", res_json, "cp932", False)

        return res_json

    # GET /views 現在のユーザーがアクセスできるすべてのビューを取得する
    def cov_get_views(self):
        """
        機能: GET /views を呼び出して現在のユーザーがアクセスできるすべてのビューを取得する
        入力: なし
        出力: get_views.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/views?locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_VIEWS
        payload = {"locale": "ja_JP"}
        print("cov_get_views:")
        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 08:06:09 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=E3316D494AFA19547B6FD1A1F98705ED; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate, no-store",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Content-Language": "ja-JP",
          "Server": "Synopsys"
        }
        """
        print(response.status_code)
        """response.status_code 失敗例 Server Error
        500
        """
        # print(response.text)
        """response.text 失敗例
        {
            "eventId": "a53908"
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("get_views.json", res_json, "shift_jis", False)

    # GET /views/user ビューの ID を取得する
    def cov_get_views_user(self):
        """
        機能: GET /views/user を呼び出してビューの ID を取得する
        入力: なし
        出力: views_user.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/views/user?locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        API_URL = self.API_BASE + self.API_VIEWS_USER
        payload = {"locale": "ja_JP"}
        print("cov_get_views_user:")
        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.headers)
        """response.headers
        {
          "Date": "Wed, 25 Jan 2023 08:54:13 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=5B544EBC2F47BA2B7968A2FC92C7C6C7; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Server": "Synopsys"
        }
        """
        print(response.status_code)
        """response.status_code
        200
        """
        # print(response.text)
        """response.text
        200
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        gljson.write_json2("views_user.json", res_json, "shift_jis", False)

    # GET /views/viewContents/{id} を使い、ビューの項目と指摘内容を取得する
    def cov_get_view_contents(self, file_path, view_id, project_id):
        """
        機能: GET /views/viewContents/{id} を呼び出して ビューの項目と指摘内容を取得する
        入力: view_id, project_id (projectKey)
        戻り値: なし
        出力: views_viewContents_view_id_project_id.json
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/views/viewContents/10551?projectId=10072&rowCount=10&offset=0&sortKey=cid&sortOrder=desc&locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        print(
            "file_path: {}, view_id: {}, project_id: {}".format(
                file_path, view_id, project_id
            )
        )

        API_URL = self.API_BASE + self.API_VIEWS_VIEWCONTENTS + "/" + view_id
        payload = {
            "locale": "ja_JP",
            "projectId": project_id,
            "rowCount": -1,
            "offset": 0,
            "sortKey": "cid",
            "sortOrder": "desc",
        }
        print("cov_get_view_contents:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )

        # print(response.headers)
        """ response.headers
        {
          "Date": "Wed, 25 Jan 2023 08:58:51 GMT",
          "Content-Type": "application/json;charset=UTF-8",
          "Transfer-Encoding": "chunked",
          "Connection": "keep-alive",
          "Set-Cookie": "COVJSESSIONID8080BV=CF9B773520E789E30553A0014DBC3334; Path=/; HttpOnly",
          "Pragma": "no-cache",
          "Cache-Control": "no-cache, no-store, must-revalidate",
          "Expires": "Thu, 01 Jan 1970 00:00:00 GMT",
          "X-Frame-Options": "DENY",
          "X-XSS-Protection": "1; mode=block",
          "X-Content-Type-Options": "nosniff",
          "vary": "accept-encoding",
          "Content-Encoding": "gzip",
          "Server": "Synopsys"
        }
      """

        # print(response.text)

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        # 拡張子なしのファイル名
        # basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        # フォルダ名（ディレクトリ名）を取得
        dirname = os.path.dirname(file_path)

        views_json_filename = (
            "views_viewContents_" + view_id + "_" + str(project_id) + ".json"
        )
        views_json_file_path = dirname + "\\" + views_json_filename
        encoding = "utf-8"
        gljson.write_json2(views_json_file_path, res_json, encoding, False)

    # POST /issues/search を使い指摘内容を取得する（「比較」でフィルターする）
    def cov_post_issues_search(
        self,
        file_path,
        projectKey,
        stream_name,
        filter_comparison,
        snapshot_scope_display,
        snapshot_scope_comparison,
    ):
        """
        機能: POST /issues/search を使い指摘内容を取得する
        入力: 保存先ファイルへのパス、プロジェクトID、ストリーム名、フィルター(比較)、スナップショットスコープ(表示)、スナップショットスコープ(比較先)、
        戻り値: JSONファイルへのパス
        出力: .json
        """
        """ curl
      # curl --location -x "http://proxy-host.example.com:3128/" \
      # --request POST "https://<coverity-host>/api/v2/issues/search? \
      # includeColumnLabels=true& \
      # locale=ja_JP& \
      # offset=0& \
      # queryType=bySnapshot& \
      # rowCount=-1& \
      # sortColumn=displayImpact& \
      # sortOrder=asc" \
      # --data "{ \
      # \"filters\": [{ \
      #   \"columnKey\": \"project\", \
      #   \"matchMode\": filter_comparison, \
      #   \"matchers\": [{\"class\": \"Project\", \"name\": project, \"type\": \"nameMatcher\"}]
      # }], \
      # \"columns\": [
      #   \"mergeKey\",
      #   \"cid\",
      #   \"displayFile\",
      #   \"fileLanguage\",
      #   \"displayFunction\",
      #   \"lineNumber\",
      #   \"displayImpact\",
      #   \"displayIssueKind\",
      #   \"displayType\",
      #   \"displayCategory\",
      #   \"cwe\",
      #   \"occurrenceCount\",
      #   \"checker\",
      #   \"status\",
      #   \"firstDetected\",
      #   \"displayComparison\",
      #   \"firstSnapshotId\",
      #   \"firstSnapshotDate\",
      #   \"firstSnapshotStream\",
      #   \"lastDetectedId\",
      #   \"lastDetected\",
      #   \"lastDetectedStream\",
      #   \"lastTriaged\",
      #   \"lastTriagedUser\",
      #   \"lastTriageComment\",
      #   \"classification\",
      #   \"severity\",
      #   \"action\",
      #   \"externalReference\",
      #   \"owner\"
      # ], \
      # \"snapshotScope\": { \
      #   \"show\": { \
      #   \"scope\": snapshot_scope_display, \
      #   \"includeOutdatedSnapshots\": false} \
      # }}" \
      # --header "Accept: application/json" \
      # --header "Content-Type: application/json" \
      # --user user2:08EED5855B845CDA35C5FD289DAA32B6
      """

        """ "matchers" の "type"
      # type には nameMatcher の他に、KeyMatcher、DateMatcher、IdMatcher がある
      here are four matcher types:
      •        idMatcher
      •        keyMatcher
      •        dateMatcher
      •        nameMatcher

      Each filter has one matcher type, except that

      NameMatcher (type: nameMatcher)
      A nameMatcher may be used in place of an idMatcher if for a filter that has both an idMatcher and a nameMatcher.
      In that case, you use a nameMatcher if you have the name of the item and not its internal identifier,
      e.g., if you have a project name instead of a project identifier.
      •        streams (class: Stream, name: stream name)
      •        project (class: Project, name: project name)
      •        owner (class: User, name: user's username)
      •        displayComponent (class: Component, name: component's name)

      IdMatcher (type: idMatcher)
      •        cid (id: cid value)
      •        cwe (id: cwe number)
      •        project (id: project id)
      •        streams (id: stream id)

      DateMatcher (type: dateMatcher)
      •        firstDetected (date: YYYY-MM-DD)
      •        lastTriaged (date: YYYY-MM-DD)
      •        firstSnapshotDate (date: YYYY-MM-DD)
      •        lastDetected (date: YYYY-MM-DD)
      Note: fetches issues with date on or after the given date.

      KeyMatcher (type: keyMatcher)
      •        checker (key: checker name)
      •        displayImpact (key: Impact value)
      •        displayCategory (key: Issue category value)
      •        displayType (key: Issue type value)
      •        displayIssueKind (key: Issue Kind value)
      •        status (key: Status value)
      •        externalReference (key: External reference url)
      •        classification (key: Classification value)
      •        severity (key: Severity value)
      •        action (key: Action value)
      •        fixTarget (key: Fix Target value)
      •        legacy (key: Legacy value)
      •        ownerFullName (key: user's full name)
      •        displayFirstDetectedBy (key: First Detected By value)
      •        displayFile (key: file name)
      •        displayFunction (key: function name)
      •        functionMergeName (key: function merge name)
      •        mergeExtra (key: merge extra)
      •        mergeKey (key: merge key)
      •        fileLanguage (key: Language value)
      •        ruleStrength (key: MISRA Category value)
      •        occurrenceCount (key: count in numerical string, fetches results with count <= given number)
      •        displayComparison (key: Comparison value)
      •        score (key: score in numerical string)
      •        All custom and standard attributes (key: attribute value)
      """

        print(
            "\nfile_path: {}\n, projectKey: {}\n, stream_name: {}\n, filter_comparison: {}\n, snapshot_scope_display: {}\n, snapshot_scope_comparison: {}\n".format(
                file_path,
                projectKey,
                stream_name,
                filter_comparison,
                snapshot_scope_display,
                snapshot_scope_comparison,
            )
        )

        API_URL = self.API_BASE + self.API_POST_ISSUES_SEARCH

        payload = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }

        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                },
                {
                    "columnKey": "displayComparison",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [{"key": filter_comparison, "type": "keyMatcher"}],
                },
            ],
            "columns": [
                "mergeKey",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "cwe",
                "occurrenceCount",
                "checker",
                "status",
                "firstDetected",
                "displayComparison",
                "firstSnapshotId",
                "firstSnapshotDate",
                "firstSnapshotStream",
                "lastDetectedId",
                "lastDetected",
                "lastDetectedStream",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
            ],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_scope_display,
                    "includeOutdatedSnapshots": False,
                },
                "compareTo": {
                    "scope": snapshot_scope_comparison,
                    "includeOutdatedSnapshots": False,
                },
            },
        }

        """ columns
      "columns": [
        "mergeKey",
        "cid",
        "displayImpact",
        "displayCategory",
        "displayFile",
        "fileLanguage",
        "displayFunction",
        "lineNumber",
        "displayImpact",
        "displayIssueKind",
        "displayType",
        "displayCategory",
        "cwe",
        "occurrenceCount",
        "checker",
        "status",
        "firstDetected",
        "displayComparison",
        "firstSnapshotId",
        "firstSnapshotDate",
        "firstSnapshotStream",
        "lastDetectedId",
        "lastDetected",
        "lastDetectedStream",
        "lastTriaged",
        "lastTriagedUser",
        "lastTriageComment",
        "classification",
        "severity",
        "action",
        "externalReference",
        "owner"
      ],
      """

        """ session 利用（不要）
      # ログインURLの設定
      url_login = "https://<coverity-host>/login/login.htm"

      # requests でログインしてセッションクッキーを作る
      session = requests.Session()    # requests.sessions.Session object at 0x7f9eefa52fd0

      # ログイン情報
      login_info = {
          "username": "user2",
          "password": "xxxxxxxx",
          # "_csrf": CSRF
      }

      # r = requests.post(url_login, data=login_info)
      # r = s.post(url_login, data=login_info, cookies=COOKIES)
      # r = session.post(url_login, data=login_info)
      """

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )

        """ response.headers
      # print(response.headers)
      {
        'Date': 'Tue, 14 Feb 2023 03:22:50 GMT',
        'Content-Type': 'application/json;charset=UTF-8',
        'Transfer-Encoding': 'chunked',
        'Connection': 'keep-alive',
        'Set-Cookie': 'COVJSESSIONID8080BV=97DD1093F7877D32259800301F23D18B; Path=/; HttpOnly',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache, no-store, must-revalidate, no-store',
        'Expires': 'Thu, 01 Jan 1970 00:00:00 GMT',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'X-Content-Type-Options': 'nosniff',
        'vary': 'accept-encoding',
        'Content-Encoding': 'gzip',
        'Content-Language': 'ja-JP',
        'Server': 'Synopsys'
      }
      """

        # print(response.text)
        # '{"statusMessage":"不正な要求","httpStatusCode":400,"detailMessage":"要求は無効です。"}'

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        # 拡張子なしのファイル名
        # basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        encoding = "utf-8"
        gljson.write_json2(file_path, res_json, encoding, False)

        return file_path

    # POST /issues/search を使い指摘内容を全数取得する（「比較」でフィルターしない）
    def cov_post_issues_search_2(
        self,
        file_path,
        projectKey,
        stream_name,
        snapshot_scope_display,
        snapshot_scope_comparison,
    ):
        """
        機能: POST /issues/search を使い指摘内容を全数取得する（「比較」でフィルターしない）
        入力: 保存先ファイルへのパス、プロジェクトID、ストリーム名、スナップショットスコープ(表示)、スナップショットスコープ(比較先)、
        戻り値: JSONファイルへのパス
        出力: .json
        """
        """ curl
      # curl --location -x "http://proxy-host.example.com:3128/" \
      # --request POST "https://<coverity-host>/api/v2/issues/search? \
      # includeColumnLabels=true& \
      # locale=ja_JP& \
      # offset=0& \
      # queryType=bySnapshot& \
      # rowCount=-1& \
      # sortColumn=displayImpact& \
      # sortOrder=asc" \
      # --data "{ \
      # \"filters\": [{ \
      #   \"columnKey\": \"project\", \
      #   \"matchMode\": filter_comparison, \
      #   \"matchers\": [{\"class\": \"Project\", \"name\": project, \"type\": \"nameMatcher\"}]
      # }], \
      # \"columns\": [
      #   \"mergeKey\",
      #   \"cid\",
      #   \"displayFile\",
      #   \"fileLanguage\",
      #   \"displayFunction\",
      #   \"lineNumber\",
      #   \"displayImpact\",
      #   \"displayIssueKind\",
      #   \"displayType\",
      #   \"displayCategory\",
      #   \"cwe\",
      #   \"occurrenceCount\",
      #   \"checker\",
      #   \"status\",
      #   \"firstDetected\",
      #   \"displayComparison\",
      #   \"firstSnapshotId\",
      #   \"firstSnapshotDate\",
      #   \"firstSnapshotStream\",
      #   \"lastDetectedId\",
      #   \"lastDetected\",
      #   \"lastDetectedStream\",
      #   \"lastTriaged\",
      #   \"lastTriagedUser\",
      #   \"lastTriageComment\",
      #   \"classification\",
      #   \"severity\",
      #   \"action\",
      #   \"externalReference\",
      #   \"owner\"
      # ], \
      # \"snapshotScope\": { \
      #   \"show\": { \
      #   \"scope\": snapshot_scope_display, \
      #   \"includeOutdatedSnapshots\": false} \
      # }}" \
      # --header "Accept: application/json" \
      # --header "Content-Type: application/json" \
      # --user user2:08EED5855B845CDA35C5FD289DAA32B6
      """

        """ "matchers" の "type"
      # type には nameMatcher の他に、KeyMatcher、DateMatcher、IdMatcher がある
      here are four matcher types:
      •        idMatcher
      •        keyMatcher
      •        dateMatcher
      •        nameMatcher

      Each filter has one matcher type, except that

      NameMatcher (type: nameMatcher)
      A nameMatcher may be used in place of an idMatcher if for a filter that has both an idMatcher and a nameMatcher.
      In that case, you use a nameMatcher if you have the name of the item and not its internal identifier,
      e.g., if you have a project name instead of a project identifier.
      •        streams (class: Stream, name: stream name)
      •        project (class: Project, name: project name)
      •        owner (class: User, name: user's username)
      •        displayComponent (class: Component, name: component's name)

      IdMatcher (type: idMatcher)
      •        cid (id: cid value)
      •        cwe (id: cwe number)
      •        project (id: project id)
      •        streams (id: stream id)

      DateMatcher (type: dateMatcher)
      •        firstDetected (date: YYYY-MM-DD)
      •        lastTriaged (date: YYYY-MM-DD)
      •        firstSnapshotDate (date: YYYY-MM-DD)
      •        lastDetected (date: YYYY-MM-DD)
      Note: fetches issues with date on or after the given date.

      KeyMatcher (type: keyMatcher)
      •        checker (key: checker name)
      •        displayImpact (key: Impact value)
      •        displayCategory (key: Issue category value)
      •        displayType (key: Issue type value)
      •        displayIssueKind (key: Issue Kind value)
      •        status (key: Status value)
      •        externalReference (key: External reference url)
      •        classification (key: Classification value)
      •        severity (key: Severity value)
      •        action (key: Action value)
      •        fixTarget (key: Fix Target value)
      •        legacy (key: Legacy value)
      •        ownerFullName (key: user's full name)
      •        displayFirstDetectedBy (key: First Detected By value)
      •        displayFile (key: file name)
      •        displayFunction (key: function name)
      •        functionMergeName (key: function merge name)
      •        mergeExtra (key: merge extra)
      •        mergeKey (key: merge key)
      •        fileLanguage (key: Language value)
      •        ruleStrength (key: MISRA Category value)
      •        occurrenceCount (key: count in numerical string, fetches results with count <= given number)
      •        displayComparison (key: Comparison value)
      •        score (key: score in numerical string)
      •        All custom and standard attributes (key: attribute value)
      """

        print(
            "\nfile_path: {}\n, projectKey: {}\n, stream_name: {}\n, snapshot_scope_display: {}\n, snapshot_scope_comparison: {}\n".format(
                file_path,
                projectKey,
                stream_name,
                snapshot_scope_display,
                snapshot_scope_comparison,
            )
        )

        API_URL = self.API_BASE + self.API_POST_ISSUES_SEARCH

        payload = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }

        # "filters" から "columnKey":"displayComparison" を削除
        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                },
            ],
            "columns": [
                "mergeKey",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "cwe",
                "occurrenceCount",
                "checker",
                "status",
                "firstDetected",
                "displayComparison",
                "firstSnapshotId",
                "firstSnapshotDate",
                "firstSnapshotStream",
                "lastDetectedId",
                "lastDetected",
                "lastDetectedStream",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
            ],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_scope_display,
                    "includeOutdatedSnapshots": False,
                },
                "compareTo": {
                    "scope": snapshot_scope_comparison,
                    "includeOutdatedSnapshots": False,
                },
            },
        }

        """ columns
      "columns": [
        "mergeKey",
        "cid",
        "displayImpact",
        "displayCategory",
        "displayFile",
        "fileLanguage",
        "displayFunction",
        "lineNumber",
        "displayImpact",
        "displayIssueKind",
        "displayType",
        "displayCategory",
        "cwe",
        "occurrenceCount",
        "checker",
        "status",
        "firstDetected",
        "displayComparison",
        "firstSnapshotId",
        "firstSnapshotDate",
        "firstSnapshotStream",
        "lastDetectedId",
        "lastDetected",
        "lastDetectedStream",
        "lastTriaged",
        "lastTriagedUser",
        "lastTriageComment",
        "classification",
        "severity",
        "action",
        "externalReference",
        "owner"
      ],
      """

        """ session 利用（不要）
      # ログインURLの設定
      url_login = "https://<coverity-host>/login/login.htm"

      # requests でログインしてセッションクッキーを作る
      session = requests.Session()    # requests.sessions.Session object at 0x7f9eefa52fd0

      # ログイン情報
      login_info = {
          "username": "user2",
          "password": "xxxxxxxx",
          # "_csrf": CSRF
      }

      # r = requests.post(url_login, data=login_info)
      # r = s.post(url_login, data=login_info, cookies=COOKIES)
      # r = session.post(url_login, data=login_info)
      """

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )

        """ response.headers
      # print(response.headers)
      {
        'Date': 'Tue, 14 Feb 2023 03:22:50 GMT',
        'Content-Type': 'application/json;charset=UTF-8',
        'Transfer-Encoding': 'chunked',
        'Connection': 'keep-alive',
        'Set-Cookie': 'COVJSESSIONID8080BV=97DD1093F7877D32259800301F23D18B; Path=/; HttpOnly',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache, no-store, must-revalidate, no-store',
        'Expires': 'Thu, 01 Jan 1970 00:00:00 GMT',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'X-Content-Type-Options': 'nosniff',
        'vary': 'accept-encoding',
        'Content-Encoding': 'gzip',
        'Content-Language': 'ja-JP',
        'Server': 'Synopsys'
      }
      """

        # print(response.text)
        # '{"statusMessage":"不正な要求","httpStatusCode":400,"detailMessage":"要求は無効です。"}'

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        # 拡張子なしのファイル名
        # basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        encoding = "utf-8"
        gljson.write_json2(file_path, res_json, encoding, False)

        return file_path

    # POST /issues/search を使い CID と mergeKey のリストを生成する
    def cov_post_issues_search_cids_mergeKeys(
        self,
        projectKey,
        stream_name,
        snapshot_id,
        output_path
    ):
        """
        目的: 
          POST /issues/search を使い CID と mergeKey のリストを生成する
        入力: 
          projectKey: プロジェクトキー
          stream_name: ストリーム名
          snapshot_id: スナップショットID
          output_path: 出力先ファイルのパス
        戻り値: 
          整形された CID と mergeKey のリスト
        出力: 
          JSON ファイル
        """

        print(
            "\nprojectKey: {},\nstream_name: {},\nsnapshot_id: {},\noutput_path: {}\n".format(
                projectKey,
                stream_name,
                snapshot_id,
                output_path
            )
        )

        payload = {
            "locale": "ja_JP",
            "rowCount": -1, # このクエリで返す行の最大数。すべての行を返すには、この値を -1 に設定します
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }
        
        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                }
            ],
            # "projectKey": projectKey,
            # "streams": [stream_name],
            "columns": ["cid", "mergeKey"],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_id,  # スナップショットIDを指定
                    "includeOutdatedSnapshots": False
                },
                "compareTo": {
                    "scope": "",
                    "includeOutdatedSnapshots": False
                }
            }
        }

        API_URL = self.API_BASE + self.API_POST_ISSUES_SEARCH

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20
        )

        # レスポンスの検証
        if response.status_code != 200:
            print(f"API call failed: {response.status_code}, {response.text}")
            return []

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        """ 異常
        {
          "offset": 0,
          "totalRows": -1,
          "columns": ["cid", "mergeKey"],
          "rows": []
        }
        """

        """ 正常
        {
          "offset": 0,
          "totalRows": 2,
          "columns": ["cid", "mergeKey"],
          "rows": [
            [{"key": "cid", "value": "564667"}, {"key": "mergeKey", "value": "key123"}],
            [{"key": "cid", "value": "564668"}, {"key": "mergeKey", "value": "key456"}]
          ]
        }
        """

        # rows から cid と mergeKey を抽出
        cid_and_mergekey_list = [
            {"cid": row[0]["value"], "mergeKey": row[1]["value"]}
            for row in res_json.get("rows", [])
            if len(row) >= 2  # cid と mergeKey があることを確認
        ]

        # # 整形したデータを保存
        encoding = "utf-8"
        with open(output_path, "w", encoding=encoding) as f:
            json.dump(cid_and_mergekey_list, f, indent=2, ensure_ascii=False)

        print(f"整形されたデータを保存しました: {output_path}")
        """
        [
          {"cid": "564667", "mergeKey": "key123"},
          {"cid": "564668", "mergeKey": "key456"}
        ]
        """

        return cid_and_mergekey_list

    # POST /issues/search を使い特定のスナップショットに関連する最初および最後のスナップショット情報を取得する
    def cov_post_issues_search_firstDetected_othres(
        self,
        projectKey,
        stream_name,
        snapshot_id,
        output_path,
    ):
        """
        目的:
          POST /issues/search を使い特定のスナップショットに関連する最初および最後のスナップショット情報を取得する
        入力:
          projectKey: プロジェクトキー
          stream_name: ストリーム名
          snapshot_id: スナップショットID
          output_path: 出力先ファイルのパス
        戻り値:
          整形された最初および最後のスナップショット情報リスト firstDetected_others_list
        出力:
          JSON ファイル
        """

        print(
            "\nprojectKey: {},\nstream_name: {},\nsnapshot_id: {},\noutput_path: {}\n".format(
                projectKey,
                stream_name,
                snapshot_id,
                output_path,
            )
        )

        payload = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }

        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                }
            ],
            # "projectKey": projectKey,
            # "streams": [stream_name],
            "columns": [
                "cid",
                "mergeKey",
                "firstDetected",            # 初回検出日
                "firstSnapshotId",          # 最初のスナップショット
                "firstSnapshotDate",        # 最初のスナップショットの日付
                "firstSnapshotStream",      # 最初のスナップショットのストリーム
                "lastDetectedId",           # 最後のスナップショット。表記の揺れあり
                "lastDetected",             # 最後のスナップショットの日付
                "lastDetectedStream",       # 最後のスナップショットのストリーム。表記の揺れあり
            ],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_id,  # スナップショットIDを指定
                    "includeOutdatedSnapshots": False
                },
                "compareTo": {
                    "scope": "",
                    "includeOutdatedSnapshots": False
                }
            }
        }

        API_URL = self.API_BASE + self.API_POST_ISSUES_SEARCH

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20
        )

        # レスポンスの検証
        if response.status_code != 200:
            print(f"API call failed: {response.status_code}, {response.text}")
            return []

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        """ 正常
        {
          'offset': 0,
          'totalRows': 47686,
          'columns': [
            'cid',
            'mergeKey',
            'firstDetected',
            'firstSnapshotId',
            'firstSnapshotDate',
            'firstSnapshotStream',
            'lastDetectedId',
            'lastDetected',
            'lastDetectedStream'
          ],
          'rows': [
            [
              {'key': 'cid', 'value': '45476'},
              {'key': 'mergeKey', 'value': '80b62a33f0a45fcadb7e7a1161862790'},
              {'key': 'firstDetected', 'value': '2017/11/22'},
              {'key': 'firstSnapshotId', 'value': '15051'},
              {'key': 'firstSnapshotDate', 'value': '2024/11/11'},
              {'key': 'firstSnapshotStream', 'value': ''},
              {'key': 'lastDetectedId', 'value': '15160'},
              {'key': 'lastDetected', 'value': '2024/12/05'},
              {'key': 'lastDetectedStream', 'value': ''}
            ],
            [...],
            ...
          ]
        }
        """
        
        # rows から firstDetectedSnapshotId 他を抽出
        firstDetected_others_list = [
            {
                "cid": row[0]["value"],
                "mergeKey": row[1]["value"],
                "firstDetected": row[2]["value"],           # 初回検出日
                "firstSnapshotId": row[3]["value"],         # 最初のスナップショット
                "firstSnapshotDate": row[4]["value"],       # 最初のスナップショットの日付
                "firstSnapshotStream": row[5]["value"],     # 最初のスナップショットのストリーム
                "lastDetectedId": row[6]["value"],          # 最後のスナップショット
                "lastDetected": row[7]["value"],            # 最後のスナップショットの日付
                "lastDetectedStream": row[8]["value"],      # 最後のスナップショットのストリーム
            }
            for row in res_json.get("rows", [])
            if len(row) >= 2  # 各 row が2つ以上の要素を持っているかを確認
        ]

        # # 整形したデータを保存
        encoding = "utf-8"
        with open(output_path, "w", encoding=encoding) as f:
            json.dump(firstDetected_others_list, f, indent=2, ensure_ascii=False)

        print(f"整形されたデータを保存しました: {output_path}")

        return firstDetected_others_list

    # POST /issueOccurrences/search を使い指摘内容を全数取得する（「比較」でフィルターする）
    def cov_post_issueOccurrences_search_2(
        self,
        file_path,
        projectKey,
        stream_name,
        filter_comparison,
        snapshot_scope_display,
        snapshot_scope_comparison,
    ):
        """
        機能: POST /issueOccurrences/search を使い指摘内容を全数取得する
        入力: 保存先ファイルへのパス、プロジェクトID、ストリーム名、フィルター(比較)、スナップショットスコープ(表示)、スナップショットスコープ(比較先)、
        戻り値: JSONファイルへのパス
        出力: .json
        """
        """ curl
      # curl --location -x "http://proxy-host.example.com:3128/" \
      # --request POST "https://<coverity-host>/api/v2/issues/search? \
      # includeColumnLabels=true& \
      # locale=ja_JP& \
      # offset=0& \
      # queryType=bySnapshot& \
      # rowCount=-1& \
      # sortColumn=displayImpact& \
      # sortOrder=asc" \
      # --data "{ \
      # \"filters\": [{ \
      #   \"columnKey\": \"project\", \
      #   \"matchMode\": filter_comparison, \
      #   \"matchers\": [{\"class\": \"Project\", \"name\": project, \"type\": \"nameMatcher\"}]
      # }], \
      # \"columns\": [
      #   \"mergeKey\",
      #   \"cid\",
      #   \"displayFile\",
      #   \"fileLanguage\",
      #   \"displayFunction\",
      #   \"lineNumber\",
      #   \"displayImpact\",
      #   \"displayIssueKind\",
      #   \"displayType\",
      #   \"displayCategory\",
      #   \"cwe\",
      #   \"occurrenceCount\",
      #   \"checker\",
      #   \"status\",
      #   \"firstDetected\",
      #   \"displayComparison\",
      #   \"firstSnapshotId\",
      #   \"firstSnapshotDate\",
      #   \"firstSnapshotStream\",
      #   \"lastDetectedId\",
      #   \"lastDetected\",
      #   \"lastDetectedStream\",
      #   \"lastTriaged\",
      #   \"lastTriagedUser\",
      #   \"lastTriageComment\",
      #   \"classification\",
      #   \"severity\",
      #   \"action\",
      #   \"externalReference\",
      #   \"owner\"
      # ], \
      # \"snapshotScope\": { \
      #   \"show\": { \
      #   \"scope\": snapshot_scope_display, \
      #   \"includeOutdatedSnapshots\": false} \
      # }}" \
      # --header "Accept: application/json" \
      # --header "Content-Type: application/json" \
      # --user user2:08EED5855B845CDA35C5FD289DAA32B6
      """

        """ "matchers" の "type"
      # type には nameMatcher の他に、KeyMatcher、DateMatcher、IdMatcher がある
      here are four matcher types:
      •        idMatcher
      •        keyMatcher
      •        dateMatcher
      •        nameMatcher

      Each filter has one matcher type, except that

      NameMatcher (type: nameMatcher)
      A nameMatcher may be used in place of an idMatcher if for a filter that has both an idMatcher and a nameMatcher.
      In that case, you use a nameMatcher if you have the name of the item and not its internal identifier,
      e.g., if you have a project name instead of a project identifier.
      •        streams (class: Stream, name: stream name)
      •        project (class: Project, name: project name)
      •        owner (class: User, name: user's username)
      •        displayComponent (class: Component, name: component's name)

      IdMatcher (type: idMatcher)
      •        cid (id: cid value)
      •        cwe (id: cwe number)
      •        project (id: project id)
      •        streams (id: stream id)

      DateMatcher (type: dateMatcher)
      •        firstDetected (date: YYYY-MM-DD)
      •        lastTriaged (date: YYYY-MM-DD)
      •        firstSnapshotDate (date: YYYY-MM-DD)
      •        lastDetected (date: YYYY-MM-DD)
      Note: fetches issues with date on or after the given date.

      KeyMatcher (type: keyMatcher)
      •        checker (key: checker name)
      •        displayImpact (key: Impact value)
      •        displayCategory (key: Issue category value)
      •        displayType (key: Issue type value)
      •        displayIssueKind (key: Issue Kind value)
      •        status (key: Status value)
      •        externalReference (key: External reference url)
      •        classification (key: Classification value)
      •        severity (key: Severity value)
      •        action (key: Action value)
      •        fixTarget (key: Fix Target value)
      •        legacy (key: Legacy value)
      •        ownerFullName (key: user's full name)
      •        displayFirstDetectedBy (key: First Detected By value)
      •        displayFile (key: file name)
      •        displayFunction (key: function name)
      •        functionMergeName (key: function merge name)
      •        mergeExtra (key: merge extra)
      •        mergeKey (key: merge key)
      •        fileLanguage (key: Language value)
      •        ruleStrength (key: MISRA Category value)
      •        occurrenceCount (key: count in numerical string, fetches results with count <= given number)
      •        displayComparison (key: Comparison value)
      •        score (key: score in numerical string)
      •        All custom and standard attributes (key: attribute value)
      """

        print(
            "\nfile_path: {}\n, projectKey: {}\n, stream_name: {}\n, snapshot_scope_display: {}\n, snapshot_scope_comparison: {}\n".format(
                file_path,
                projectKey,
                stream_name,
                snapshot_scope_display,
                snapshot_scope_comparison,
            )
        )

        # Request URL
        API_URL = self.API_BASE + self.API_POST_ISSUEOCCURRENCES_SEARCH

        payload = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }

        # "filters" から "columnKey":"displayComparison" を削除
        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                },
                {
                    "columnKey": "displayComparison",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [{"key": filter_comparison, "type": "keyMatcher"}],
                },
            ],
            "columns": [
                "mergeKey",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "cwe",
                "occurrenceCount",
                "checker",
                "status",
                "firstDetected",
                "displayComparison",
                "firstSnapshotId",
                "firstSnapshotDate",
                "firstSnapshotStream",
                "lastDetectedId",
                "lastDetected",
                "lastDetectedStream",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
            ],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_scope_display,
                    "includeOutdatedSnapshots": False,
                },
                "compareTo": {
                    "scope": snapshot_scope_comparison,
                    "includeOutdatedSnapshots": False,
                },
            },
        }

        """ columns
      "columns": [
        "mergeKey",
        "cid",
        "displayImpact",
        "displayCategory",
        "displayFile",
        "fileLanguage",
        "displayFunction",
        "lineNumber",
        "displayImpact",
        "displayIssueKind",
        "displayType",
        "displayCategory",
        "cwe",
        "occurrenceCount",
        "checker",
        "status",
        "firstDetected",
        "displayComparison",
        "firstSnapshotId",
        "firstSnapshotDate",
        "firstSnapshotStream",
        "lastDetectedId",
        "lastDetected",
        "lastDetectedStream",
        "lastTriaged",
        "lastTriagedUser",
        "lastTriageComment",
        "classification",
        "severity",
        "action",
        "externalReference",
        "owner"
      ],
      """

        """ session 利用（不要）
      # ログインURLの設定
      url_login = "https://<coverity-host>/login/login.htm"

      # requests でログインしてセッションクッキーを作る
      session = requests.Session()    # requests.sessions.Session object at 0x7f9eefa52fd0

      # ログイン情報
      login_info = {
          "username": "user2",
          "password": "xxxxxxxx",
          # "_csrf": CSRF
      }

      # r = requests.post(url_login, data=login_info)
      # r = s.post(url_login, data=login_info, cookies=COOKIES)
      # r = session.post(url_login, data=login_info)
      """

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )

        if response.status_code != 200:
            print(
                "[cov_post_issueOccurrences_search_2] Stream doesnt exist: ",
                response.status_code,
            )

            sys.exit(604)

        """ response.headers
      # print(response.headers)
      {
        'Date': 'Tue, 14 Feb 2023 03:22:50 GMT',
        'Content-Type': 'application/json;charset=UTF-8',
        'Transfer-Encoding': 'chunked',
        'Connection': 'keep-alive',
        'Set-Cookie': 'COVJSESSIONID8080BV=97DD1093F7877D32259800301F23D18B; Path=/; HttpOnly',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache, no-store, must-revalidate, no-store',
        'Expires': 'Thu, 01 Jan 1970 00:00:00 GMT',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'X-Content-Type-Options': 'nosniff',
        'vary': 'accept-encoding',
        'Content-Encoding': 'gzip',
        'Content-Language': 'ja-JP',
        'Server': 'Synopsys'
      }
      """

        # print(response.text)
        # '{"statusMessage":"不正な要求","httpStatusCode":400,"detailMessage":"要求は無効です。"}'

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        # 拡張子なしのファイル名
        # basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        encoding = "utf-8"
        gljson.write_json2(file_path, res_json, encoding, False)

        return file_path

    # POST /issueOccurrences/search を使い指摘内容を全数取得する（「比較」でフィルターしない）
    def cov_post_issueOccurrences_search(
        self,
        file_path,
        projectKey,
        stream_name,
        snapshot_scope_display,
        snapshot_scope_comparison,
    ):
        """
        機能: POST /issueOccurrences/search を使い指摘内容を全数取得する
        入力: 保存先ファイルへのパス、プロジェクトID、ストリーム名、スナップショットスコープ(表示)、スナップショットスコープ(比較先)、
        戻り値: JSONファイルへのパス
        出力: .json
        """
        """ curl
      # curl --location -x "http://proxy-host.example.com:3128/" \
      # --request POST "https://<coverity-host>/api/v2/issueOccurrences/search? \
      # includeColumnLabels=true& \
      # locale=ja_JP& \
      # offset=0& \
      # queryType=bySnapshot& \
      # rowCount=-1& \
      # sortColumn=displayImpact& \
      # sortOrder=asc" \
      # --data "{ \
      # \"filters\": [{ \
      #   \"columnKey\": \"project\", \
      #   \"matchMode\": filter_comparison, \
      #   \"matchers\": [{\"class\": \"Project\", \"name\": project, \"type\": \"nameMatcher\"}]
      # }], \
      # \"columns\": [
      #   \"mergeKey\",
      #   \"cid\",
      #   \"displayFile\",
      #   \"fileLanguage\",
      #   \"displayFunction\",
      #   \"lineNumber\",
      #   \"displayImpact\",
      #   \"displayIssueKind\",
      #   \"displayType\",
      #   \"displayCategory\",
      #   \"cwe\",
      #   \"occurrenceCount\",
      #   \"checker\",
      #   \"status\",
      #   \"firstDetected\",
      #   \"displayComparison\",
      #   \"firstSnapshotId\",
      #   \"firstSnapshotDate\",
      #   \"firstSnapshotStream\",
      #   \"lastDetectedId\",
      #   \"lastDetected\",
      #   \"lastDetectedStream\",
      #   \"lastTriaged\",
      #   \"lastTriagedUser\",
      #   \"lastTriageComment\",
      #   \"classification\",
      #   \"severity\",
      #   \"action\",
      #   \"externalReference\",
      #   \"owner\"
      # ], \
      # \"snapshotScope\": { \
      #   \"show\": { \
      #   \"scope\": snapshot_scope_display, \
      #   \"includeOutdatedSnapshots\": false} \
      # }}" \
      # --header "Accept: application/json" \
      # --header "Content-Type: application/json" \
      # --user user2:08EED5855B845CDA35C5FD289DAA32B6
      """

        """ "matchers" の "type"
      # type には nameMatcher の他に、KeyMatcher、DateMatcher、IdMatcher がある
      here are four matcher types:
      •        idMatcher
      •        keyMatcher
      •        dateMatcher
      •        nameMatcher

      Each filter has one matcher type, except that

      NameMatcher (type: nameMatcher)
      A nameMatcher may be used in place of an idMatcher if for a filter that has both an idMatcher and a nameMatcher.
      In that case, you use a nameMatcher if you have the name of the item and not its internal identifier,
      e.g., if you have a project name instead of a project identifier.
      •        streams (class: Stream, name: stream name)
      •        project (class: Project, name: project name)
      •        owner (class: User, name: user's username)
      •        displayComponent (class: Component, name: component's name)

      IdMatcher (type: idMatcher)
      •        cid (id: cid value)
      •        cwe (id: cwe number)
      •        project (id: project id)
      •        streams (id: stream id)

      DateMatcher (type: dateMatcher)
      •        firstDetected (date: YYYY-MM-DD)
      •        lastTriaged (date: YYYY-MM-DD)
      •        firstSnapshotDate (date: YYYY-MM-DD)
      •        lastDetected (date: YYYY-MM-DD)
      Note: fetches issues with date on or after the given date.

      KeyMatcher (type: keyMatcher)
      •        checker (key: checker name)
      •        displayImpact (key: Impact value)
      •        displayCategory (key: Issue category value)
      •        displayType (key: Issue type value)
      •        displayIssueKind (key: Issue Kind value)
      •        status (key: Status value)
      •        externalReference (key: External reference url)
      •        classification (key: Classification value)
      •        severity (key: Severity value)
      •        action (key: Action value)
      •        fixTarget (key: Fix Target value)
      •        legacy (key: Legacy value)
      •        ownerFullName (key: user's full name)
      •        displayFirstDetectedBy (key: First Detected By value)
      •        displayFile (key: file name)
      •        displayFunction (key: function name)
      •        functionMergeName (key: function merge name)
      •        mergeExtra (key: merge extra)
      •        mergeKey (key: merge key)
      •        fileLanguage (key: Language value)
      •        ruleStrength (key: MISRA Category value)
      •        occurrenceCount (key: count in numerical string, fetches results with count <= given number)
      •        displayComparison (key: Comparison value)
      •        score (key: score in numerical string)
      •        All custom and standard attributes (key: attribute value)
      """

        print(
            "\nfile_path: {}\n, projectKey: {}\n, stream_name: {}\n, snapshot_scope_display: {}\n, snapshot_scope_comparison: {}\n".format(
                file_path,
                projectKey,
                stream_name,
                snapshot_scope_display,
                snapshot_scope_comparison,
            )
        )

        # Request URL
        API_URL = self.API_BASE + self.API_POST_ISSUEOCCURRENCES_SEARCH

        payload = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "sortOrder": "asc",
            "includeColumnLabels": True,
            "queryType": "bySnapshot",
            "sortColumn": "displayImpact",
        }

        # "filters" から "columnKey":"displayComparison" を削除
        data_dict = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ],
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": stream_name, "type": "nameMatcher"}
                    ],
                },
            ],
            "columns": [
                "mergeKey",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "cwe",
                "occurrenceCount",
                "checker",
                "status",
                "firstDetected",
                "displayComparison",
                "firstSnapshotId",
                "firstSnapshotDate",
                "firstSnapshotStream",
                "lastDetectedId",
                "lastDetected",
                "lastDetectedStream",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
            ],
            "snapshotScope": {
                "show": {
                    "scope": snapshot_scope_display,
                    "includeOutdatedSnapshots": False,
                },
                "compareTo": {
                    "scope": snapshot_scope_comparison,
                    "includeOutdatedSnapshots": False,
                },
            },
        }

        """ columns
      "columns": [
        "mergeKey",
        "cid",
        "displayImpact",
        "displayCategory",
        "displayFile",
        "fileLanguage",
        "displayFunction",
        "lineNumber",
        "displayImpact",
        "displayIssueKind",
        "displayType",
        "displayCategory",
        "cwe",
        "occurrenceCount",
        "checker",
        "status",
        "firstDetected",
        "displayComparison",
        "firstSnapshotId",
        "firstSnapshotDate",
        "firstSnapshotStream",
        "lastDetectedId",
        "lastDetected",
        "lastDetectedStream",
        "lastTriaged",
        "lastTriagedUser",
        "lastTriageComment",
        "classification",
        "severity",
        "action",
        "externalReference",
        "owner"
      ],
      """

        """ session 利用（不要）
      # ログインURLの設定
      url_login = "https://<coverity-host>/login/login.htm"

      # requests でログインしてセッションクッキーを作る
      session = requests.Session()    # requests.sessions.Session object at 0x7f9eefa52fd0

      # ログイン情報
      login_info = {
          "username": "user2",
          "password": "xxxxxxxx",
          # "_csrf": CSRF
      }

      # r = requests.post(url_login, data=login_info)
      # r = s.post(url_login, data=login_info, cookies=COOKIES)
      # r = session.post(url_login, data=login_info)
      """

        # POST
        response = requests.post(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            json=data_dict,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )

        """ response.headers
      # print(response.headers)
      {
        'Date': 'Tue, 14 Feb 2023 03:22:50 GMT',
        'Content-Type': 'application/json;charset=UTF-8',
        'Transfer-Encoding': 'chunked',
        'Connection': 'keep-alive',
        'Set-Cookie': 'COVJSESSIONID8080BV=97DD1093F7877D32259800301F23D18B; Path=/; HttpOnly',
        'Pragma': 'no-cache',
        'Cache-Control': 'no-cache, no-store, must-revalidate, no-store',
        'Expires': 'Thu, 01 Jan 1970 00:00:00 GMT',
        'X-Frame-Options': 'DENY',
        'X-XSS-Protection': '1; mode=block',
        'X-Content-Type-Options': 'nosniff',
        'vary': 'accept-encoding',
        'Content-Encoding': 'gzip',
        'Content-Language': 'ja-JP',
        'Server': 'Synopsys'
      }
      """

        # print(response.text)
        # '{"statusMessage":"不正な要求","httpStatusCode":400,"detailMessage":"要求は無効です。"}'

        # JSON 整形
        res_json = response.json()
        # print(json.dumps(res_json, indent=4))

        # GLJson インスタンス化
        gljson = GLJson()

        # JSONファイル保存
        # 拡張子なしのファイル名
        # basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        encoding = "utf-8"
        gljson.write_json2(file_path, res_json, encoding, False)

        return file_path

    # JSONファイルを読み込み pandas DataFrame に変換する
    def convert_json_df(self, f, encoding):
        df_f = pd.read_json(f, encoding=encoding)

        # All arrays must be of the same length

        print("df_f: {}".format(df_f.head(10)))

        return df_f

    # リストを読み込み pandas DataFrame に変換する
    def convert_list_df(self, l):
        df_l = pd.DataFrame(l)
        print("df_l: {}".format(df_l.head(10)))

        return df_l

    # 複雑な要素
    def cov_json_normalize(self, l):
        df_l = pd.json_normalize(
            l,
            # record_path='key',
            # meta=['key', 'value'],
        )
        print(df_l.head(10))

    # JSONファイルを Pandas DataFrame に変換して CSVファイルに保存
    def convert_json_df_csv(self, file_path, encoding):
        """
        機能: JSONファイルを Pandas DataFrame に変換して CSVファイルに保存する
        入力: ディレクトリパスを含むJSONファイル（例: "./dir/subdir/issues.json"）、エンコーディングの種類（例: "shift-jis"）
        出力: 変換後のCSVファイル（拡張子のみ変更）
        戻り値: 変換後のデータフレーム
        """

        # GLJson インスタンス化
        # インスタンス化しないと self.read_json_2() で 'COVApi' object has no attribute 'read_json_2'
        gljson = GLJson()

        # JSONファイルの読み込み（要素１つのリストの中に辞書）
        err, json_file_list = gljson.read_json_2(file_path, encoding)

        if err != 0:
            return err

        # リストを pandas DataFrame に変換
        df = self.convert_list_df(json_file_list)

        # 拡張子なしのファイル名
        basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        # フォルダ名（ディレクトリ名）を取得
        dirname = os.path.dirname(file_path)

        # CSVファイル保存（CSVファイルは cp932 で保存）
        csv_file_path = dirname + "\\" + basename_without_ext + ".csv"
        encoding = "cp932"
        df.to_csv(csv_file_path, header=True, index=False, encoding=encoding)

        return df

    # views_viewContents.json ファイルから "rows" を pandas DataFrame に変換する（独自ループ）
    # 13,000件に3分以上かかるので使わない
    def convert_list_row_df(self, json_file, encoding):
        # GLJson インスタンス化
        # インスタンス化しないと self.read_json_2() で 'COVApi' object has no attribute 'read_json_2'
        gljson = GLJson()

        # *.json 読み込み
        err, views_viewContents_dict = gljson.read_json_2(json_file, encoding)

        # 指摘件数
        views_viewContents_list_len = len(views_viewContents_dict["rows"])

        # 要素 "rows" 抽出
        views_viewContents_rows_list = views_viewContents_dict["rows"]
        print(views_viewContents_rows_list)

        # インデックス[0] のみ（テスト）
        index_0_df = pd.json_normalize(views_viewContents_rows_list[0])
        print(index_0_df)

        # 転置
        """転置前
                                 key                                              value
        0                        cid                                             430811
        1                displayFile  /cov/groups/d-net/sample_projectA/...
        2               fileLanguage                                                  C
        3            displayFunction                                  EvSche_InitModule
        ... snip ...
        39                  mergeKey                   db67f17a460d15fdff078efcf1a31150
        40         functionMergeName                                  EvSche_InitModule
        41                     score       
        """

        """転置後
                   0                                                  1             2   ...                                39                 40     41
        key       cid                                        displayFile  fileLanguage  ...                          mergeKey  functionMergeName  score 
        value  430811  /cov/groups/d-net/sample_projectA/...             C  ...  db67f17a460d15fdff078efcf1a31150  EvSche_InitModule        
        """
        index_0_T_df = index_0_df.T
        print(index_0_T_df)

        # 初期化
        df_viewContents = pd.DataFrame(data=[], index=[], columns=[])

        """indexと異なり、columnsはカラム名の設定にはならないので注意
        df_viewContents.columns = [
                "index",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "checker",
                "cwe",
                "occurrenceCount",
                "status",
                "firstDetected",
                "displayComparison",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
                "fixTarget",
                "lastDetectedId",
                "lastDetectedStream",
                "lastDetectedTarget",
                "lastDetectedVersion",
                "lastDetected",
                "lastDetectedDescription",
                "firstSnapshotId",
                "firstSnapshotStream",
                "firstSnapshotDate",
                "firstSnapshotTarget",
                "firstSnapshotVersion",
                "firstSnapshotDescription",
                "displayComponent",
                "legacy",
                "ownerFullName",
                "mergeKey",
                "functionMergeName",
                "score",
        ]
        
        df_viewContents = pd.DataFrame(
            data=[],
            index=[], 
            columns=[
                "index",
                "cid",
                "displayFile",
                "fileLanguage",
                "displayFunction",
                "lineNumber",
                "displayImpact",
                "displayIssueKind",
                "displayType",
                "displayCategory",
                "checker",
                "cwe",
                "occurrenceCount",
                "status",
                "firstDetected",
                "displayComparison",
                "lastTriaged",
                "lastTriagedUser",
                "lastTriageComment",
                "classification",
                "severity",
                "action",
                "externalReference",
                "owner",
                "fixTarget",
                "lastDetectedId",
                "lastDetectedStream",
                "lastDetectedTarget",
                "lastDetectedVersion",
                "lastDetected",
                "lastDetectedDescription",
                "firstSnapshotId",
                "firstSnapshotStream",
                "firstSnapshotDate",
                "firstSnapshotTarget",
                "firstSnapshotVersion",
                "firstSnapshotDescription",
                "displayComponent",
                "legacy",
                "ownerFullName",
                "mergeKey",
                "functionMergeName",
                "score",
            ]
        )
        """

        print(df_viewContents)

        # 指摘件数分回す
        for x in range(views_viewContents_list_len):
            df_x = pd.json_normalize(views_viewContents_rows_list[x])
            # print(df_x.T[1:2])
            df_viewContents = pd.concat(
                [df_viewContents, df_x.T[1:2]], axis=0, ignore_index=True
            )

        print(df_viewContents)

        # columns リネーム
        df_viewContents_new = df_viewContents.rename(
            columns={0: "cid", 39: "mergeKey"},
        )

        print(df_viewContents_new)

        # CSV 保存
        df_viewContents_new.to_csv(
            "df_viewContents_loop.csv", header=True, index=False, encoding=encoding
        )

        return df_viewContents_new

    # 辞書からデータフレームを作成（from_dict）
    def convert_dic_df_from_dict(self, json_file, encoding):
        """変換前 dic_rows
        "rows": [
            [
                {
                    "key": "cid",
                    "value": "430811"
                },
                {
                    "key": "displayFile",
                    "value": "/cov/groups/d-net/sample_projectA/....c"
                },
                ... snip ...
                {
                    "key": "fileLanguage",
                    "value": "C"
                },
                {
                    "key": "displayFunction",
                    "value": "EvSche_InitModule"
                },
                {
                    "key": "lineNumber",
                    "value": "74"
                },
                {
                    "key": "displayImpact",
                    "value": "Low"
                },
                {
                    "key": "displayIssueKind",
                    "value": "Quality"
                },
                {
                    "key": "displayType",
                    "value": "CERT-C Characters and Strings"
                },
                {
                    "key": "displayCategory",
                    "value": "Coding standard violation"
                },
                {
                    "key": "checker",
                    "value": "CERT STR30-C"
                },
                {
                    "key": "cwe",
                    "value": "758"
                },
                {
                    "key": "occurrenceCount",
                    "value": "1"
                },
                {
                    "key": "status",
                    "value": "New"
                },
                {
                    "key": "firstDetected",
                    "value": "01/05/23"
                },
                {
                    "key": "displayComparison",
                    "value": "Absent"
                },
                {
                    "key": "lastTriaged",
                    "value": ""
                },
                {
                    "key": "lastTriagedUser",
                    "value": ""
                },
                {
                    "key": "lastTriageComment",
                    "value": ""
                },
                {
                    "key": "classification",
                    "value": "Unclassified"
                },
                {
                    "key": "severity",
                    "value": "Unspecified"
                },
                {
                    "key": "action",
                    "value": "Undecided"
                },
                {
                    "key": "externalReference",
                    "value": ""
                },
                {
                    "key": "owner",
                    "value": "Unassigned"
                },
                {
                    "key": "fixTarget",
                    "value": "Untargeted"
                },
                {
                    "key": "lastDetectedId",
                    "value": "14002"
                },
                {
                    "key": "lastDetectedStream",
                    "value": "sample_projectA_coverity"
                },
                {
                    "key": "lastDetectedTarget",
                    "value": ""
                },
                {
                    "key": "lastDetectedVersion",
                    "value": ""
                },
                {
                    "key": "lastDetected",
                    "value": "01/11/23"
                },
                {
                    "key": "lastDetectedDescription",
                    "value": ""
                },
                {
                    "key": "firstSnapshotId",
                    "value": "14000"
                },
                {
                    "key": "firstSnapshotStream",
                    "value": "sample_projectA_coverity"
                },
                {
                    "key": "firstSnapshotDate",
                    "value": "01/10/23"
                },
                {
                    "key": "firstSnapshotTarget",
                    "value": ""
                },
                {
                    "key": "firstSnapshotVersion",
                    "value": ""
                },
                {
                    "key": "firstSnapshotDescription",
                    "value": ""
                },
                {
                    "key": "displayComponent",
                    "value": "Other"
                },
                {
                    "key": "legacy",
                    "value": "False"
                },
                {
                    "key": "ownerFullName",
                    "value": ""
                },
                {
                    "key": "mergeKey",
                    "value": "db67f17a460d15fdff078efcf1a31150"
                },
                {
                    "key": "functionMergeName",
                    "value": "EvSche_InitModule"
                },
                {
                    "key": "score",
                    "value": ""
                }
            ],
            ... snip ...
            [
                ... snip ...
            ]
        ]
        """

        with open(json_file, "r", encoding=encoding) as f:
            s = json.load(f)

        # 文字列を辞書やリストからなるオブジェクトに変換する
        # d = json.loads(s["rows"])
        rows_l = s["rows"]

        new_l = list(
            map(lambda tags: {tag["key"]: tag["value"] for tag in tags}, rows_l)
        )
        df = pd.DataFrame.from_dict(new_l)

        """ 変換後
        ,cid,1,2,3,4,5,6,7,8,9,10,...,35,36,37,mergeKey,39,40,41
        0, '430811', '/cov/groups/d-net/sample_projectA/.../EvSche.c', ... snip ... , ''}"

        """

        # すべての欠損値 'NaN'を '-' に置き換える（不要）
        # df.fillna("-")
        print(df.head(10))

        # CSV 保存
        df.to_csv(
            json_file + "_dic_df_from_dict.csv",
            header=True,
            index=False,
            encoding=encoding,
        )

        return df

    # 辞書からデータフレームを作成（内包表記）
    def convert_dict_df_comprehensions(self, file_path, encoding):
        """
        機能: 辞書を Pandas DataFrame に変換して CSVファイルに CP932（固定）で保存する
        入力: ディレクトリパスを含むJSONファイル（例: "./dir/subdir/views_viewContents.json"）、エンコーディングの種類（例: "shift-jis"）
        出力: 変換後のCSVファイル（拡張子のみ変更）
        戻り値: 変換後のデータフレーム, パスを含むCSVファイル名
        """

        # JSONファイル読み込み (UTF-8)
        with open(file_path, "r", encoding=encoding) as f:
            s = json.load(f)

        try:
            # 文字列を辞書やリストからなるオブジェクトに変換する
            # d = json.loads(s["rows"])
            rows_l = s["rows"]

        except KeyError as err:
            print("[convert_dict_df_comprehensions] KeyError!: ", err)

            sys.exit(600)

        # 新しい辞書を作成
        new_l = [{tag["key"]: tag["value"] for tag in tags} for tags in rows_l]

        # 変換
        df = pd.DataFrame.from_dict(new_l)

        # print("df.isnull().sum(): {}".format(df.isnull().sum()))
        # print("df.isnull().values: {}".format(df.isnull().values))
        # print("df.head(10): ".format(df.head(10)))
        # すべての欠損値 'NaN'を '-' に置き換える
        # df.fillna("-")
        print(df.head(10))

        # カラム名を Coverity Connect 画面に合わせる（日本語に変更）
        # 英語
        new_columns_En = {
            "cid": "CID",
            "displayFile": "File",
            "fileLanguage": "Language",
            "displayFunction": "Function",
            "lineNumber": "Line Number",
            "displayImpact": "Impact",
            "displayIssueKind": "Issue Kind",
            "displayType": "Type",
            "displayCategory": "Category",
            "checker": "Checker",
            "cwe": "CWE",
            "occurrenceCount": "Count",
            "status": "Status",
            "firstDetected": "First Detected",
            "displayComparison": "Comparison",
            "lastTriaged": "Last Triaged",
            "lastTriagedUser": "Last Triaged User",
            "lastTriageComment": "Last Triage Comment",
            "classification": "Classification",
            "severity": "Severity",
            "action": "Action",
            "externalReference": "External Reference",
            "owner": "Owner",
            "fixTarget": "Fix Target",
            "lastDetectedId": "Last Snapshot",
            "lastDetectedStream": "Last Snapshot Stream",
            "lastDetectedTarget": "Last Snapshot Target",
            "lastDetectedVersion": "Last Snapshot Version",
            "lastDetected": "Last Snapshot Date",
            "lastDetectedDescription": "Last Snapshot Description",
            "firstSnapshotId": "First Snapshot",
            "firstSnapshotStream": "First Snapshot Stream",
            "firstSnapshotDate": "First Snapshot Date",
            "firstSnapshotTarget": "First Snapshot Target",
            "firstSnapshotVersion": "First Snapshot Version",
            "firstSnapshotDescription": "First Snapshot Description",
            "displayComponent": "Component",
            "legacy": "Legacy",
            "ownerFullName": "Owner Name",
            "mergeKey": "Merge Key",
            "functionMergeName": "Function Merge Name",
            "score": "Score",
        }

        # 日本語
        new_columns_JP = {
            "cid": "CID",
            "displayFile": "ファイル",
            "fileLanguage": "言語",
            "displayFunction": "関数名",
            "lineNumber": "行番号",
            "displayImpact": "影響度",
            "displayIssueKind": "問題の種類",
            "displayType": "型",
            "displayCategory": "カテゴリ",
            "checker": "チェッカー",
            "cwe": "CWE",
            "occurrenceCount": "カウント",
            "status": "状態",
            "firstDetected": "初回検出日(First Detected)",
            "displayComparison": "比較(Comparison)",
            "lastTriaged": "最終選別日(Last Triaged)",
            "lastTriagedUser": "最終選別ユーザー",
            "lastTriageComment": "最終選別コメント",
            "classification": "分類",
            "severity": "重要度",
            "action": "アクション",
            "externalReference": "外部参照",
            "owner": "所有者/担当者",
            "fixTarget": "修正対象",
            "lastDetectedId": "最後のスナップショット",
            "lastDetectedStream": "最後のスナップショットのストリーム",
            "lastDetectedTarget": "最後のスナップショットのターゲット",
            "lastDetectedVersion": "最後のスナップショットのバージョン",
            "lastDetected": "最後のスナップショットの日付(Last Snapshot Date)",
            "lastDetectedDescription": "最後のスナップショットの説明",
            "firstSnapshotId": "最初のスナップショット",
            "firstSnapshotStream": "最初のスナップショットのストリーム",
            "firstSnapshotDate": "最初のスナップショットの日付(First Snapshot Date)",
            "firstSnapshotTarget": "最初のスナップショットのターゲット",
            "firstSnapshotVersion": "最初のスナップショットのバージョン",
            "firstSnapshotDescription": "最初のスナップショットの説明",
            "displayComponent": "コンポーネント",
            "legacy": "レガシー",
            "ownerFullName": "担当者名",
            "mergeKey": "mergeKey",
            "functionMergeName": "関数結合名(Function Merge Name)",
            "score": "スコア",
        }

        # カラム名変更
        # 直接、元のオブジェクトを変更するときは、
        # df_viewContents_3.rename(columns=new_columns, inplace=True)
        new_df = df.rename(columns=new_columns_JP)
        print(new_df.head(10))

        # 拡張子なしのファイル名
        basename_without_ext = os.path.splitext(os.path.basename(file_path))[0]

        # フォルダ名（ディレクトリ名）を取得
        dirname = os.path.dirname(file_path)

        # CSVファイル保存 (CP932)
        csv_file_path = dirname + "\\" + basename_without_ext + ".csv"
        encoding = "cp932"
        new_df.to_csv(csv_file_path, header=True, index=False, encoding=encoding)

        return new_df, csv_file_path

    # 内部結合
    def inner_join(
        self,
        df_left,
        df_right,
        file_dir,
        prefix,
        latest,
        latestBefore,
        filter_comparison,
        encoding,
    ):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を内部結合する
        入力: データフレーム左、データフレーム右、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、接頭語（new, fixed, unfixed）、
         表示欄に入力するスナップショットID、比較先に入力するスナップショットID、比較、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"inner_join.csv"）
        戻り値: 結合後のデータフレーム、結合したCSVファイルへのパス
        """
        # 保存するファイルへのパス
        file_path = (
            file_dir
            + prefix
            + "_inner_join-"
            + latest
            + "_"
            + latestBefore
            + "_"
            + filter_comparison
            + ".csv"
        )

        # 内部結合
        df = pd.merge(
            df_left, df_right, how="inner", on="mergeKey", indicator=True, sort=True
        )

        # CSVファイル保存
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df, file_path

    # 左外部結合
    def left_outer_join(self, df_left, df_right, file_dir, encoding):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を左外部結合する
        入力: データフレーム1、データフレーム2、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"left_outer_join.csv"）
        戻り値: 結合後のデータフレーム
        """
        # 保存するファイルへのパス
        file_path = file_dir + "left_outer_join.csv"

        # 左外部結合
        df = pd.merge(df_left, df_right, how="left", on="mergeKey")

        # CSVファイル保存
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df

    # 左外部結合（ファイル名にプリフィックス追加）
    def left_outer_join_2(
        self,
        df_left,
        df_right,
        file_dir,
        prefix,
        latest,
        latestBefore,
        filter_comparison,
        encoding,
    ):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を左外部結合する
        入力: データフレーム1、データフレーム2、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、接頭語（new, fixed, unfixed）、
         表示欄に入力するスナップショットID、比較先に入力するスナップショットID、比較、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"left_outer_join.csv"）
        戻り値: 結合後のデータフレーム
        """
        # 保存するファイルへのパス
        # file_path = file_dir + "left_outer_join.csv"
        file_path = (
            file_dir
            + prefix
            + "_left_outer_join-"
            + latest
            + "_"
            + latestBefore
            + "_"
            + filter_comparison
            + ".csv"
        )

        # 左外部結合
        df = pd.merge(df_left, df_right, how="left", on="mergeKey")

        # CSVファイル保存
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df

    # 右外部結合
    def right_outer_join(self, df_left, df_right, file_dir, encoding):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を右外部結合する
        入力: データフレーム1、データフレーム2、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"right_outer_join.csv"）
        戻り値: 結合後のデータフレーム、結合したCSVファイルへのパス
        """
        # 保存するファイルへのパス
        file_path = (
            file_dir
            + "right_outer_join-"
            + self.group
            + "-"
            + self.project
            + "-"
            + self.branch
            + "-"
            + self.commit_short
            + ".csv"
        )

        # 右外部結合
        df = pd.merge(df_left, df_right, how="right", on="mergeKey")

        # 列順を変更
        print("（前）df.columns: {}".format(df.columns))
        new_order = [
            "CID",
            "ファイル",
            "言語",
            "関数名",
            "functionDisplayName",
            "行番号",
            "影響度",
            "問題の種類",
            "型",
            "type2",
            "カテゴリ",
            "category2",
            "eventDescription",
            "CWE",
            "カウント",
            "チェッカー",
            "状態",
            "初回検出日(First Detected)",
            "比較(Comparison)",
            "最初のスナップショット",
            "最初のスナップショットの日付(First Snapshot Date)",
            "最初のスナップショットのストリーム",
            "最後のスナップショット",
            "最後のスナップショットの日付(Last Snapshot Date)",
            "最後のスナップショットのストリーム",
            "最終選別日(Last Triaged)",
            "最終選別ユーザー",
            "最終選別コメント",
            "分類",
            "重要度",
            "アクション",
            "外部参照",
            "所有者/担当者",
            "mergeKey",
        ]

        df = df[new_order]
        print("（後）df.columns: {}".format(df.columns))

        # CSVファイル保存
        # エクセルでファイルを開いているとエラーが出る
        # PermissionError: [Errno 13] Permission denied: \
        # 'C:\\cov\\groups\\wlan-adapter\\sample_project\\cov_issues\\right_outer_join.csv'
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df, file_path

    # 右外部結合（ファイル名にプリフィックス追加）
    def right_outer_join_2(
        self,
        df_left,
        df_right,
        file_dir,
        prefix,
        latest,
        latestBefore,
        filter_comparison,
        encoding,
    ):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を右外部結合する
        入力: データフレーム左、データフレーム右、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、接頭語（new, fixed, unfixed）、
         表示欄に入力するスナップショットID、比較先に入力するスナップショットID、比較、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"right_outer_join.csv"）
        戻り値: 結合後のデータフレーム、結合したCSVファイルへのパス
        """
        # 保存するファイルへのパス
        # file_path = file_dir + prefix + "_right_outer_join-" + self.group + "-" +self.project + "-" + self.branch +"-" +self.commit_short + ".csv"
        file_path = (
            file_dir
            + prefix
            + "_right_outer_join-"
            + latest
            + "_"
            + latestBefore
            + "_"
            + filter_comparison
            + ".csv"
        )

        # 右外部結合
        # 引数indicatorをTrueとし、_mergeという列が追加され、both, left_only, right_onlyのいずれかに分類される
        # 引数sortをTrueとし、キー列でソートする
        df = pd.merge(
            df_left, df_right, how="right", on="mergeKey", indicator=True, sort=True
        )

        # 列順を変更
        print("（前）df.columns: {}".format(df.columns))
        new_order = [
            "CID",
            "ファイル",
            "言語",
            "関数名",
            "functionDisplayName",
            "行番号",
            "影響度",
            "問題の種類",
            "型",
            "type2",
            "カテゴリ",
            "category2",
            "eventDescription",
            "CWE",
            "カウント",
            "チェッカー",
            "状態",
            "初回検出日(First Detected)",
            "比較(Comparison)",
            "最初のスナップショット",
            "最初のスナップショットの日付(First Snapshot Date)",
            "最初のスナップショットのストリーム",
            "最後のスナップショット",
            "最後のスナップショットの日付(Last Snapshot Date)",
            "最後のスナップショットのストリーム",
            "最終選別日(Last Triaged)",
            "最終選別ユーザー",
            "最終選別コメント",
            "分類",
            "重要度",
            "アクション",
            "外部参照",
            "所有者/担当者",
            "mergeKey",
            "_merge",
        ]

        df = df[new_order]
        print("（後）df.columns: {}".format(df.columns))

        # CSVファイル保存
        # エクセルでファイルを開いているとエラーが出る
        # PermissionError: [Errno 13] Permission denied: \
        # 'C:\\cov\\groups\\wlan-adapter\\sample_project\\cov_issues\\right_outer_join.csv'
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df, file_path

    # 右外部結合 CSVファイル用
    def right_outer_join_2_csv(
        self,
        df_left,
        df_right,
        file_dir,
        prefix,
        latest,
        latestBefore,
        filter_comparison,
        encoding,
    ):
        """
        機能: 列 "CID" をキーに、2つの pandas.DataFrame を右外部結合する（CSVファイル用）
        入力:
          df_left: 左側の DataFrame
          df_right: 右側の DataFrame
          file_dir: CSVファイル保存先ディレクトリ（例: "./dir/subdir"）
          prefix: 接頭語（例: "new", "fixed", "unfixed"）
          latest: 表示用の直近スナップショットID
          latestBefore: 比較先の直近より前のスナップショットID
          filter_comparison: 比較フィルター（例: "PRESENT", "ABSENT"）
          encoding: エンコーディング（例: "cp932"）
        出力:
          結合後の CSV ファイル
        戻り値:
          結合後の DataFrame と、結合した CSV ファイルへのパス
        """
        # 保存するファイルへのパスの生成
        file_path = (
            file_dir
            + prefix
            + "_right_outer_join-"
            + latest
            + "_"
            + latestBefore
            + "_"
            + filter_comparison
            + ".csv"
        )
    
        # 'CID' 列を同じ型（文字列型）に統一
        df_left.loc[:, 'CID'] = df_left['CID'].astype(str)
        df_right.loc[:, 'CID'] = df_right['CID'].astype(str)
    
        # 右外部結合（indicator=True で結合状況を確認）
        df = pd.merge(
            df_left, df_right, how="right", on="CID", indicator=True, sort=True
        )
    
        # 結合結果のカラム並び順を調整
        new_order = [
            "CID",
            "ファイル",
            "言語",
            "関数名",
            "functionDisplayName",
            "行番号",
            "影響度",
            "問題の種類",
            "型",
            "type2",
            "カテゴリ",
            "category2",
            "eventDescription",
            "CWE",
            "カウント",
            "チェッカー",
            "状態",
            "初回検出日(First Detected)",
            "比較(Comparison)",
            "最初のスナップショット",
            "最初のスナップショットの日付(First Snapshot Date)",
            "最初のスナップショットのストリーム",
            "最後のスナップショット",
            "最後のスナップショットの日付(Last Snapshot Date)",
            "最後のスナップショットのストリーム",
            "最終選別日(Last Triaged)",
            "最終選別ユーザー",
            "最終選別コメント",
            "分類",
            "重要度",
            "アクション",
            "外部参照",
            "所有者/担当者",
            "_merge",
        ]
        # new_order に含まれているカラムのみを残す
        new_order = [col for col in new_order if col in df.columns]
        df = df[new_order]
    
        # CSV ファイルとして保存（エンコーディングは cp932）
        df.to_csv(file_path, header=True, index=False, encoding=encoding)
    
        return df, file_path

    # 完全外部結合
    def full_outer_join(self, df_left, df_right, file_dir, encoding):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を完全外部結合する
        入力: データフレーム1、データフレーム2、CSVファイル保存先ディレクトリ（例: "./dir/subdir"）、エンコーディングの種類（例: "shift-jis"）
        出力: 結合後のCSVファイル（"full_outer_join.csv"）
        戻り値: 結合後のデータフレーム
        """
        # 保存するファイルへのパス
        file_path = file_dir + "full_outer_join.csv"

        # 完全外部結合
        df = pd.merge(df_left, df_right, how="outer", on="mergeKey")

        # CSVファイル保存
        df.to_csv(file_path, header=True, index=False, encoding=encoding)

        return df

    # データフレームの結合4種
    def join_all(self, df1, df2, file_dir, encoding):
        """
        機能: 列 "mergeKey" をキーに、2つの pandas.DataFrame を4種類の方法で結合する
        入力: データフレーム1、データフレーム2、CSVファイル保存先ディレクトリ（"./dir/subdir"）、
        CSVファイル用エンコーディングの種類（"shift-jis"）
        出力: 結合後のCSVファイル（"inner_join.csv" 他）
        戻り値: なし
        """
        # 内部結合
        df = self.inner_join(df1, df2, file_dir, encoding)
        print("INNER CONNECT: ".format(df.head(10)))

        # 左外部結合
        df = self.left_outer_join(df1, df2, file_dir, encoding)
        print("LEFT OUTER CONNECT: ".format(df.head(10)))

        # 右外部結合
        df = self.right_outer_join(df1, df2, file_dir, encoding)
        print("RIGHT OUTER CONNECT: ".format(df.head(10)))

        # 完全外部結合
        df = self.full_outer_join(df1, df2, file_dir, encoding)
        print("FULL OUTER CONNECT: ".format(df.head(10)))

    # 認定ユーザー検証
    # 環境変数から秘密情報を取得する
    def get_env_variable(self, key):
        """
        機能: 秘密情報を環境変数から取得する
        入力（引数）: key
        出力（戻り値）: key の値
        使用例: 認証キーを取得
        cov_auth_key = get_env_variable('COVAUTHKEY')
        """
        err = 0
        try:
            return os.environ[key]

        except KeyError as e:
            print(f"Error: Environment variable '{key}' not found. {e}")
            err = 708
            sys.exit(err)

    # 認定ユーザーの判定（ファイル利用）
    def user_id_exists_in_file(self, file_path, user_id):
        """
        機能: 利用者のユーザーIDが認定ユーザーであるか否かを判定する
        入力（引数）: ファイルパス、ユーザーID
        出力（戻り値）: 認定ユーザーの場合 True、認定ユーザーでない場合 False
        使用例:
        file_path = 'path/to/your/csvfile.csv'  # CSVファイルのパス
        user_id = 'user2'  # 検索したいユーザーID
        user_exists = user_id_exists_in_file(file_path, user_id)
        """
        """ ファイルの書式（Coverity Connect > 設定 > ユーザーおよびグループ > ユーザー・タブ > エクスポート で出力するCSVファイル）
        ユーザーID、名、姓、メールアドレス、組込みユーザーフラグ、所属グループ、最終ログイン日時
        "reporter","Reporter","User","","true","ユーザー;","2021-11-01T06:17:45.667Z"
        "user1","太郎","山田","user1@example.com","false","ユーザー;","2020-07-17T06:40:20.134Z"
        "user2","花子","佐藤","user2@example.com","false","ユーザー;アドミニストレータ;","2021-11-01T06:17:34.737Z"
         :
        """
        with open(file_path, mode="r", encoding="utf-8") as file:
            reader = csv.reader(file)
            # next(reader)  # ヘッダー行がある場合スキップ
            for row in reader:
                if row[0] == user_id:
                    return True

            return False

    # 電子メールアドレスからユーザーIDを取得する（ファイル利用）
    def get_user_id_by_email(self, file_path, email):
        """
        機能: 利用者の電子メールアドレスからユーザーIDを取得する
        入力（引数）: ファイルパス、電子メールアドレス
        出力（戻り値）: 認定ユーザーの場合 ユーザーID、認定ユーザーでない場合 None
        使用例:
        file_path = 'path/to/your/csvfile.csv'  # CSVファイルのパス
        email = 'user2@example.com'  # 検索したい電子メールアドレス
        user_id = get_user_id_by_email(file_path, email)
        """
        """ ファイルの書式（Coverity Connect > 設定 > ユーザーおよびグループ > ユーザー・タブ > エクスポート で出力するCSVファイル）
        ユーザーID、名、姓、メールアドレス、組込みユーザーフラグ、所属グループ、最終ログイン日時
        "reporter","Reporter","User","","true","ユーザー;","2021-11-01T06:17:45.667Z"
        "user1","太郎","山田","user1@example.com","false","ユーザー;","2020-07-17T06:40:20.134Z"
        "user2","花子","佐藤","user2@example.com","false","ユーザー;アドミニストレータ;","2021-11-01T06:17:34.737Z"
         :
        """
        with open(file_path, mode="r", encoding="utf-8") as file:
            reader = csv.reader(file)
            for row in reader:
                if row[3] == email:  # メールアドレスは4番目の列（インデックス3）にある
                    return row[0]  # ユーザーIDを返す

            return None

    # GET Retrieve user を使いユーザーIDが認定ユーザーか判定する（API利用）
    def user_id_exists(self, user_id):
        """
        機能: GET Retrieve user を使いユーザーIDが認定ユーザーか判定する
        入力（引数）: ユーザーID
        出力（戻り値）: 認定ユーザーの場合 True、認定ユーザーでない場合 False
        使用例:
        user_id = 'user2'  # 検索したいユーザーID
        user_info = user_id_exists(self, user_id)
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request GET "https://<coverity-host>/api/v2/users/user2?locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6

        API_URL = self.API_BASE + self.api_userS + "/" + user_id
        payload = {
            # "locale": "ja_JP",
        }
        print("user_id_exists:")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }

        {"code":1100,"message":"hasuike に一致するユーザーが見つかりません。"}
        """
        """response 成功（アドミニストレーター）
        {
            "users": [
                {
                  "createdBy": "admin",
                  "dateCreated": "2017-05-20T01:40:31.411Z",
                  "dateDeleted": null,
                  "dateModified": "2023-11-27T03:58:31.590Z",
                  "deletedBy": null,
                  "disabled": false,
                  "domainName": null,
                  "email": "user2@example.com",
                  "familyName": "xxxx",
                  "given_name": "yyyy",
                  "groupNames": [
                    "Administrators",
                    "Users"
                  ],
                  "lastLogin": "2023-11-27T03:58:31.589Z",
                  "local": true,
                  "locale": "ja_JP",
                  "locked": false,
                  "modifiedBy": "system",
                  "name": "user2",
                  "passwordChanged": "2021-09-27T16:34:11.159Z",
                  "roleAssignments": [
                    {
                      "group": null,
                      "roleAssignmentType": "user",
                      "roleName": "projectOwner",
                      "scope": "project",
                      "username": "user2"
                    },
                    ... snip ...
                    {
                      "group": null,
                      "roleAssignmentType": "user",
                      "roleName": "projectOwner",
                      "scope": "project",
                      "username": "user2"
                    }
                  ],
                  "superUser": false
                }
            ]
        }
        """
        """ response 成功（一般ユーザー）
        {
            "users": [
                {
                  "createdBy": "admin",
                  "dateCreated": "2019-12-18T01:43:45.190Z",
                  "dateDeleted": null,
                  "dateModified": "2023-05-23T06:45:39.241Z",
                  "deletedBy": null,
                  "disabled": false,
                  "domainName": null,
                  "email": "user3@example.com",
                  "familyName": "xxxx",
                  "givenName": "yyyy",
                  "groupNames": [
                    "Users"
                  ],
                  "lastLogin": "2023-05-23T06:45:39.235Z",
                  "local": true,
                  "locale": "ja_JP",
                  "locked": false,
                  "modifiedBy": "user3",
                  "name": "user3",
                  "passwordChanged": "2021-08-05T15:42:06.743Z",
                  "roleAssignments": [
                    {
                      "group": null,
                      "roleAssignmentType": "user",
                      "roleName": "developer",
                      "scope": "global",
                      "username": "user3"
                    }
                  ],
                  "superUser": false
                }
            ]
        }
        """

        # JSON 整形
        res_json = response.json()
        print(json.dumps(res_json, indent=4))

        # users > disabled 抽出
        try:
            users_disabled = res_json["users"][0]["disabled"]
            print("users > disabled: {}".format(users_disabled))

            if not users_disabled:
                # 認定ユーザーの場合 users_disabled は False
                return True

            else:
                # 認定ユーザーでない場合 users_disabled は True
                return False

        except KeyError as err:
            print("[user_id_exists] KeyError!: ", err)
            # err = 708
            return False

    # Get Retrieve all users を使い有効なすべてのユーザ情報を取得する（API利用）
    def cov_get_all_users(self):
        """
        機能: GET Retrieve all users を使い有効なすべてのユーザ―情報を取得する
        入力（引数）: なし
        出力（戻り値）: 有効なすべてのユーザー情報
        """
        # curl --location -x "http://proxy-host.example.com:3128/" ^
        # --request GET "https://<coverity-host>/api/v2/users?disabled=false&includeDetails=true&locked=false&offset=0&rowCount=200&sortColumn=name&sortOrder=asc" ^
        # --header "Accept: application/json" ^
        # --header "Content-Type: application/json" ^
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6

        API_URL = self.API_BASE + self.api_userS
        payload = {
            "disabled": "false",
            "includeDetails": "true",
            "locked": "false",
            "offset": "0",
            "rowCount": "200",
            "sortColumn": "name",
            "sortOrder": "asc",
        }
        print("[cov_get_all_users] Now processing.")

        response = requests.get(
            API_URL,
            proxies=REQUESTS_PROXIES,
            headers=self.API_HEADERS,
            params=payload,
            auth=self.API_AUTH,
            verify=False,
            timeout=20,
        )
        # print(response.text)
        """response.text エラーの例
        {
          "statusMessage": "見つかりません",
          "httpStatusCode": 404,
          "detailMessage": "申し訳ありませんが、アクセスしようとしているリンクはどこにも行きません。"
        }

        {"users":[]}
        """

        """ response 成功 2023/11/27
        {
        "users": [
        {
          "createdBy": null,
          "dateCreated": "2017-03-22T03:22:40.026802Z",
          "dateDeleted": null,
          "dateModified": "2023-07-10T14:10:28.327Z",
          "deletedBy": null,
          "disabled": false,
          "domainName": null,
          "email": "user2@example.com",
          "familyName": "User",
          "givenName": "Admin",
          "groupNames": [
            "Administrators",
            "Users"
          ],
          "lastLogin": "2023-07-10T14:10:28.325Z",
          "local": true,
          "locale": "ja_JP",
          "locked": false,
          "modifiedBy": "admin",
          "name": "admin",
          "passwordChanged": "2023-07-10T14:09:56.716Z",
          "roleAssignments": [
            {
              "group": null,
              "roleAssignmentType": "user",
              "roleName": "streamOwner",
              "scope": "stream",
              "username": "admin"
            },
            ... snip ...
            {
              "group": null,
              "roleAssignmentType": "user",
              "roleName": "streamOwner",
              "scope": "stream",
              "username": "admin"
            }
          ],
          "superUser": true
        },
        {
          "createdBy": null,
          "dateCreated": "2022-11-30T17:46:27Z",
          "dateDeleted": null,
          "dateModified": "2023-09-29T03:15:02.946Z",
          "deletedBy": null,
          "disabled": false,
          "domainName": null,
          "email": "user4@example.com",
          "familyName": "xxxx",
          "givenName": "yyyy",
          "groupNames": [
            "Users"
          ],
          "lastLogin": "2023-01-17T00:56:21.813Z",
          "local": true,
          "locale": "ja_JP",
          "locked": false,
          "modifiedBy": "system",
          "name": "user4",
          "passwordChanged": "2023-01-16T20:42:33.574Z",
          "roleAssignments": [],
          "superUser": false
        },
        ... snip ...
        {
          "createdBy": null,
          "dateCreated": "2023-03-22T07:35:53.742Z",
          "dateDeleted": null,
          "dateModified": "2023-10-24T12:28:12.236Z",
          "deletedBy": null,
          "disabled": false,
          "domainName": null,
          "email": "tools@example.com",
          "familyName": "xxxx",
          "givenName": "yyyy",
          "groupNames": [
            "Users"
          ],
          "lastLogin": "2023-10-24T12:28:12.232Z",
          "local": true,
          "locale": "en_US",
          "locked": false,
          "modifiedBy": "yyyy",
          "name": "yyyy",
          "passwordChanged": "2023-03-22T09:14:51.698Z",
          "roleAssignments": [],
          "superUser": false
        }
        ]
        }
        """

        # users の要素数チェック
        try:
            # JSON 整形
            res_json = response.json()
            len_users = len(res_json["users"])  # users 要素数
            print("length of users: {}".format(len_users))

            # 認定ユーザー authorized_users をファイルに保存する
            # file_path = "S\shared\coverity\authorized_user\authorized_user.json"

            # COVProj クラスと同様の方法で OS 判定
            # COVProj のインスタンスがある場合
            # file_path = os.path.join(covproj.share_dir, "authorized_user", "authorized_user.json")
            if os.name == "nt":
            	# Windows
                file_path = r"S:\\shared\\coverity\\authorized_user\\authorized_user.json"

            else:
                # Linux
                file_path = "/mnt/share/coverity/authorized_user/authorized_user.json"

            # res_jsonをJSON形式の文字列に変換します
            authorized_users = json.dumps(res_json, indent=4)

            # ファイルを書き込みモードで開き、JSONデータを書き込みます
            with open(file_path, "w") as file:
                file.write(authorized_users)

            return res_json

        except Exception as err:
            print("[cov_get_all_users] Exception!: ", err)
            err = {"Keyerror": 708}

            return err

    # 利用者（の電子メール）が認定ユーザーの場合、ユーザーIDを得る
    def check_user_id(self, email):
        """
        機能: 利用者（の電子メール）が認定ユーザーか判定する
        入力（引数）: 電子メールアドレス
        出力（戻り値）: 認定ユーザーの場合 ユーザーID、認定ユーザーでない場合 None
        使用例:
        email = 'user2@example.com'  # 検索したい電子メールアドレス
        user_id = check_user_id(self, email)
        """
        # 電子メールアドレスを検索
        # self.get_user_id_by_email(file_path, email)

        print("[check_user_id] Now processing.")

        # 電子メールが認定ユーザーかチェックする
        all_users = self.cov_get_all_users()

        for user in all_users["users"]:
            if user["email"] == email:
                # user_id を返す
                return user["name"]

            else:
                pass

        return None

    # パスワード生成器
    def generate_password(self, length):
        # 使用する文字のセットを定義
        characters = string.ascii_letters + string.digits + string.punctuation
        # ランダムに文字を選んでパスワードを生成
        password = "".join(secrets.choice(characters) for i in range(length))
        return password

    # アカウント発行メール送信
    def send_account_issuance_email(self, user_id, user_email, generated_password):
        """
        機能: アカウント発行メールを送信する
        入力（引数）: ユーザーID、ユーザー・メールアドレス、パスワード
        出力: メール
        """
        # SMTPサーバーの設定
        smtp_server = os.getenv("COVAUTO_SMTP_HOST", "localhost")
        smtp_port = 25
        # smtp_user = "user2@example.com"
        # smtp_password = "your_password"

        # user_id = "tester"
        # user_email = "user2@example.com"
        # generated_password = "12345"

        # メールの設定
        from_email = "CovLicense <user2@example.com>"
        to_email = user_email
        cc_emails = [
            "user2@example.com",
            "user5@example.com",
            "user6@example.com",
            "user7@example.com",
            "user3@example.com",
        ]
        # cc_emails = ["4m9y2c@bma.biglobe.ne.jp"]
        subject = f"[Coverity] アカウント発行連絡 ({user_id})"
        body = f"""

        ※ このメールは、システムから自動送信されています。

        ユーザー {user_id} を、Coverity Connect サーバーに登録しました。
        仮パスワードは、 {generated_password} です。

        Coverity Connect サーバー https://<coverity-host>/ にログインし、
        仮パスワードを変更してください。

        以上
        """

        # メールの作成
        msg = MIMEMultipart()
        msg["From"] = from_email
        msg["To"] = to_email
        msg["Cc"] = ", ".join(cc_emails)
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        # 送信先リストにCCのアドレスを追加
        recipients = [to_email] + cc_emails
        print(f"recipients: {recipients}")

        # メールの送信
        try:
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            # server.login(smtp_user, smtp_password)
            server.sendmail(from_email, recipients, msg.as_string())
            server.quit()
            print("Email sent successfully")

            return 0

        except Exception as e:
            print(f"Failed to send email: {e}")

            return 1

    # 認定ユーザーの改廃（新規登録、登録解除）
    # POST Create new user を使い新規ユーザーを登録する
    def create_new_user(self, user_info):
        """
        機能: 新規ユーザーを登録する
        入力（引数）: ユーザー情報
        出力（戻り値）: 成功 201 / 失敗 400, 401, 403, 409, 500, 例外理由
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request POST "https://<coverity-host>/api/v2/users?locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        # -d {...}

        email = user_info["email"]
        family_name = user_info["family_name"]
        given_name = user_info["given_name"]
        user_id = user_info["user_id"]

        API_URL = self.API_BASE + self.api_userS
        payload = {"locale": "ja_JP"}

        # パスワードの長さ
        password_length = 8
        # パスワードを生成
        generated_password = self.generate_password(password_length)
        print(f"Generated Password: {generated_password}")

        data_dict = {
            "disabled": False,
            # "domainName": "local",
            "email": email,
            "familyName": family_name,
            "givenName": given_name,
            # "groups": [
            #  {
            #    "domainName": "local",
            #    "name": user_id
            #  }
            # ],
            # "local": "true",
            "locale": "ja_JP",
            # "locked": "false",
            "name": user_id,
            # "password": "coverity",
            "password": generated_password,
            # "roleAssignments": [
            #  {
            #    "group": {
            #      "domainName": "local",
            #      "name": "Users"
            #    },
            #    "roleAssignmentType": "user",
            #    "roleName": "developer",
            #    "scope": "global",
            #    "username": user_id
            #  }
            # ]
        }

        # POST
        try:
            response = requests.post(
                API_URL,
                proxies=REQUESTS_PROXIES,
                headers=self.API_HEADERS,
                params=payload,
                json=data_dict,
                auth=self.API_AUTH,
                verify=False,
                timeout=20,
            )

            # JSON 整形
            print(f"response.status_code: {response.status_code}")
            if response.status_code == 201:
                # ユーザー作成に成功
                # アカウント発行メール送信
                self.send_account_issuance_email(user_id, email, generated_password)

                return 201

            else:
                # ユーザー作成に失敗
                res_json = response.json()
                print(json.dumps(res_json, indent=4))

                ret_code = res_json["code"]
                ret_message = res_json["message"]
                print(f"code: {ret_code}, message: {ret_message}")

                return response.status_code

        except Timeout as err:
            print(f"[create_new_user] タイムアウトエラーが発生しました: {err}")
            return str(err)

        except ConnectionError as err:
            print(f"[create_new_user] 接続エラーが発生しました: {err}")
            return str(err)

        except HTTPError as err:
            print(f"[create_new_user] HTTPエラーが発生しました: {err}")
            return str(err)

        except TooManyRedirects as err:
            print(f"[create_new_user] リダイレクトが多すぎます: {err}")
            return str(err)

        except Exception as err:
            print(f"[create_new_user] Exception: : {err}")
            return str(err)

    # PUT Update user を使いユーザー情報を更新する（未使用）
    def update_user(self, user_info):
        """
        機能: ユーザー情報を更新する
        入力（引数）: ユーザーID
        出力（戻り値）: 成功 True / 失敗 False
        """
        pass

    # DELETE Delete user を使いユーザーを削除する
    def delete_user(self, user_id):
        """
        機能: ユーザーを削除する
        入力（引数）: ユーザーID
        出力（戻り値）: 成功 200 / 失敗 400, 401, 403, 409, 500, 例外理由
        """
        # curl --location -x "http://proxy-host.example.com:3128/" \
        # --request POST "https://<coverity-host>/api/v2/users?locale=ja_JP" \
        # --header "Accept: application/json" \
        # --header "Content-Type: application/json" \
        # --user user2:08EED5855B845CDA35C5FD289DAA32B6
        # -d {...}

        # email = user_info["email"]
        # family_name = user_info["familyName"]
        # given_name = user_info["givenName"]
        # user_id = user_info["user_id"]

        API_URL = self.API_BASE + self.api_userS + "/" + user_id
        payload = {
            "locale": "ja_JP"
        }

        # DELETE
        try:
            response = requests.delete(
                API_URL,
                proxies=REQUESTS_PROXIES,
                headers=self.API_HEADERS,
                params=payload,
                # json=data_dict,
                auth=self.API_AUTH,
                verify=False,
                timeout=20,
            )

            # JSON 整形
            print(f"response.status_code: {response.status_code}")
            if response.status_code == 200:
                # ユーザー削除に成功
                return 200

            else:
                # ユーザー削除に失敗
                res_json = response.json()
                print(json.dumps(res_json, indent=4))

                ret_code = res_json["code"]
                ret_message = res_json["message"]
                print(f"code: {ret_code}, message: {ret_message}")

                return response.status_code

        except Timeout as err:
            print(f"[delete_user] タイムアウトエラーが発生しました: {err}")
            return str(err)

        except ConnectionError as err:
            print(f"[delete_user] 接続エラーが発生しました: {err}")
            return str(err)

        except HTTPError as err:
            print(f"[delete_user] HTTPエラーが発生しました: {err}")
            return str(err)

        except TooManyRedirects as err:
            print(f"[delete_user] リダイレクトが多すぎます: {err}")
            return str(err)

        except Exception as err:
            print(f"[delete_user] Exception: : {err}")
            return str(err)

    # チェッカー vs CWE
    # チェッカー取得、CWE付与（指定したスナップショットのみ）
    def get_checker_cwe_mapping(
        self,
        projectKey: int,
        streamName: str,
        snapshotScopeShow: int,
        snapshotScopeCompare: int = None,
        proxies: dict = None,
        timeout: float = 30.0
    ):
        """
        チェッカー名と CWE ID の対応表を生成し、CSV を出力する。
        チェッカー ⇔ CWE の対応表を生成し、CSV を出力する。
        ただし、指定スナップショットに実際に問題を報告したチェッカーのみ

        :param projectKey: int, プロジェクト ID
        :param streamName: str, ストリーム名
        :param snapshotScopeShow: int, 対象スナップショット ID
        :param snapshotScopeCompare: int or None, 比較先スナップショット ID
        :param proxies: dict or None, requests 用プロキシ設定
        :param timeout: float, リクエストタイムアウト秒
        """
        url = self.API_BASE + self.API_POST_ISSUES_SEARCH

        # パラメータ：全件取得、ソート不要、ラベル不要
        # クエリパラメータ
        params = {
            "locale": "ja_JP",
            "rowCount": -1,
            "offset": 0,
            "queryType": "bySnapshot",
            "includeColumnLabels": False,
            "sortOrder": "asc"
        }
        # フィルターとスナップショット範囲を設定
        body = {
            "filters": [
                {
                    "columnKey": "project",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Project", "id": projectKey, "type": "idMatcher"}
                    ]
                },
                {
                    "columnKey": "streams",
                    "matchMode": "oneOrMoreMatch",
                    "matchers": [
                        {"class": "Stream", "name": streamName, "type": "nameMatcher"}
                    ]
                }
            ],
            "columns": ["checker", "cwe"],
            "snapshotScope": {
                "show": {
                    "scope": snapshotScopeShow,
                    "includeOutdatedSnapshots": False
                },
                "compareTo": {
                    "scope": snapshotScopeCompare or "",
                    "includeOutdatedSnapshots": False
                }
            }
        }

        session = requests.Session()
        # trust_env=False にすると環境変数プロキシ設定を無視しますが、
        # explicit proxies パラメータは利用されます。
        session.trust_env = False

        resp = requests.post(
            url,
            params=params,
            json=body,
            headers=self.API_HEADERS,
            auth=self.API_AUTH,
            verify=False,
            proxies=proxies,
            timeout=timeout
        )
        resp.raise_for_status()
        data = resp.json()

        # JSON の rows から {checker:…, cwe:…} を抽出
        records = []
        for row in data.get("rows", []):
            entry = {col["key"]: col["value"] for col in row}
            records.append(entry)

        # DataFrame 化 → 重複除去
        df = pd.DataFrame(records)
        df = df.drop_duplicates().reset_index(drop=True)

        # 1) CSV出力
        df.to_csv("checker_cwe_mapping.csv", index=False, encoding="utf-8-sig")

        # 2) Markdown 出力（tabulate が無い場合はスキップ）
        # try:
        #     with open("checker_cwe_mapping.md", "w", encoding="utf-8") as f:
        #         f.write(df.to_markdown(index=False))
        # except ImportError:
        #     print("警告: 'tabulate' が見つからないため、Markdown 出力をスキップしました。")

        return df

    # 全チェッカーを取得するメソッド
    def get_all_checkers(self, proxies=None, timeout=30):
        url = self.API_BASE + "/checkerAttributes/checker"
        session = requests.Session(); session.trust_env = False
        resp = session.get(
            url,
            headers=self.API_HEADERS,
            auth=self.API_AUTH,
            verify=False,
            proxies=proxies,
            timeout=timeout
        )
        resp.raise_for_status()
        data = resp.json()
        
        # レスポンス例: {"checkerAttribute":{...},"checkerAttributedata":[{"key":"BAD_FREE","value":"Bad free"},…]}
        return [item["key"] for item in data["checkerAttributedata"]]

